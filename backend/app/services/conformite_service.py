"""
Service d'analyse de conformité contractuelle.
Compare la FPP (version DB) contre les Conditions Générales uploadées (PDF/Word).
Analyse article par article avec plusieurs engagements par article.
"""
import io, json

MAX_CG_CHARS = 18000


# ── Extraction texte CG ───────────────────────────────────────────────────────

def extract_text_from_document(file_bytes: bytes, filename: str) -> str:
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext == "pdf":
        return _extract_pdf(file_bytes)
    elif ext in ("docx", "doc"):
        return _extract_word(file_bytes)
    try:
        return file_bytes.decode("utf-8", errors="replace")
    except Exception:
        raise ValueError(f"Format non supporté : {filename}. Utilisez PDF ou Word.")


def _extract_pdf(file_bytes: bytes) -> str:
    from pdfminer.high_level import extract_text
    try:
        return (extract_text(io.BytesIO(file_bytes)) or "").strip()
    except Exception as e:
        raise ValueError(f"Erreur lecture PDF : {e}")


def _extract_word(file_bytes: bytes) -> str:
    from docx import Document
    try:
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        raise ValueError(f"Erreur lecture Word : {e}")


# ── Fetch FPP depuis DB ───────────────────────────────────────────────────────

def get_fpp_rows(db, product_id: int, fpp_version: int) -> list[dict]:
    from app.models.fiche_direct import FicheDirectItem
    items = (
        db.query(FicheDirectItem)
        .filter(
            FicheDirectItem.product_id == product_id,
            FicheDirectItem.version_number == fpp_version,
        )
        .order_by(FicheDirectItem.sheet, FicheDirectItem.section, FicheDirectItem.id)
        .all()
    )
    return [
        {
            "parametre": i.parameter,
            "valeur": i.value or "Information manquante",
            "section": i.section or "",
            "sheet": i.sheet or "",
        }
        for i in items
    ]


# ── Analyse principale ────────────────────────────────────────────────────────

def analyze_conformite(
    db,
    product_id: int,
    fpp_version: int,
    cg_bytes: bytes,
    cg_filename: str,
    provider: str = "openai-gpt5",
) -> dict:
    fpp_rows = get_fpp_rows(db, product_id, fpp_version)
    if not fpp_rows:
        raise ValueError(f"Aucun paramètre FPP trouvé pour le produit {product_id} version {fpp_version}.")

    cg_text = extract_text_from_document(cg_bytes, cg_filename)
    if len(cg_text) > MAX_CG_CHARS:
        cg_text = cg_text[:MAX_CG_CHARS] + "\n[... texte tronqué ...]"

    fpp_summary = [
        {"parametre": r["parametre"], "valeur": r["valeur"], "section": r["section"]}
        for r in fpp_rows[:100]
    ]

    from app.services.ai_service import _call, set_active_provider, _parse_json
    set_active_provider(provider)

    prompt = f"""Tu es un expert senior en conformité contractuelle assurance-vie collective Art. 83 (retraite d'entreprise).
Tu maîtrises parfaitement la réglementation assurance (Code des assurances, Loi Fillon, ANI, directive Solvabilité II), les calculs actuariels, les règles de gestion et la migration de systèmes d'information assurantiels.

MISSION : Analyser les Conditions Générales (CG) article par article et vérifier que chaque engagement contractuel est correctement pris en compte dans la Fiche Produit Paramétré (FPP).

=== FPP — Paramétrage en base (version {fpp_version}) ===
{json.dumps(fpp_summary, ensure_ascii=False)}

=== CONDITIONS GÉNÉRALES ===
{cg_text}

INSTRUCTIONS D'ANALYSE :

1. IDENTIFIER tous les articles et sections des CG.
2. Pour CHAQUE article, extraire TOUS les engagements contractuels distincts (un article peut contenir plusieurs engagements).
3. Pour CHAQUE engagement :
   a. Chercher sa correspondance dans la FPP
   b. Attribuer un STATUT :
      - "conforme"            : engagement correctement paramétré dans la FPP
      - "partiel"             : engagement partiellement couvert (valeur approchée ou incomplète)
      - "non_repris"          : engagement absent de la FPP — à paramétrer
      - "validation_requise"  : présent mais nécessite une validation métier ou actuarielle
      - "sans_impact"         : engagement sans impact direct sur le paramétrage SI
   c. Identifier le TYPE DE RISQUE si statut ≠ conforme/sans_impact :
      - "juridique" : risque de contestation contractuelle ou légale
      - "actuariel" : impact sur calcul de provisions, primes, rentes
      - "financier" : impact sur frais, rendements, reversements
      - "operationnel" : risque de traitement incorrect des opérations
      - "reglementaire" : non-conformité réglementaire (Code assurances, ANI, etc.)
   d. NIVEAU DE RISQUE : "critique" | "eleve" | "moyen" | "faible"
   e. CITATION exacte du texte CG concerné (max 2 phrases)
   f. RÉFÉRENCE FPP : le paramètre FPP correspondant et sa valeur (ou null si absent)
   g. EXPLICATION : pourquoi c'est important + conséquences si non paramétré (1-2 phrases)

Retourne UNIQUEMENT ce JSON valide :
{{
  "score_conformite": <entier 0-100>,
  "resume": "<synthèse 2-3 phrases>",
  "articles": [
    {{
      "article": "<référence article ex: Art. 12 — Frais de gestion>",
      "engagements": [
        {{
          "engagement": "<description de l'engagement contractuel>",
          "citation_cg": "<extrait exact du texte CG>",
          "statut": "conforme|partiel|non_repris|validation_requise|sans_impact",
          "risque_type": "juridique|actuariel|financier|operationnel|reglementaire|null",
          "risque_niveau": "critique|eleve|moyen|faible",
          "reference_fpp": "<parametre FPP = valeur, ou null>",
          "explication": "<explication métier + conséquences>"
        }}
      ]
    }}
  ]
}}"""

    try:
        raw = _call("gpt-4o", prompt, max_tokens=16000)
        result = _parse_json(raw)
    except Exception as e:
        raise ValueError(f"Erreur analyse LLM : {e}")

    articles = result.get("articles", [])
    statuses = ("conforme", "partiel", "non_repris", "validation_requise", "sans_impact")
    summary = {s: 0 for s in statuses}
    total = 0
    for art in articles:
        for eng in art.get("engagements", []):
            s = eng.get("statut", "sans_impact")
            if s in summary:
                summary[s] += 1
            total += 1

    return {
        "score_conformite": result.get("score_conformite", 0),
        "resume": result.get("resume", ""),
        "cg_filename": cg_filename,
        "fpp_version": fpp_version,
        "fpp_params": len(fpp_rows),
        "total_engagements": total,
        "articles": articles,
        "summary": summary,
    }


# ── Historisation ─────────────────────────────────────────────────────────────

def save_conformite_history(db, product_id: int, cg_filename: str, fpp_version: int, provider: str, result: dict) -> int:
    import json as _json
    from app.models.conformite_history import ConformiteHistory

    entry = ConformiteHistory(
        product_id=product_id,
        filename_kelia=f"FPP v{fpp_version}",
        filename_contract=cg_filename,
        provider=provider,
        kelia_params=result.get("fpp_params", 0),
        score_conformite=result.get("score_conformite", 0),
        result_json=_json.dumps(result, ensure_ascii=False),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry.id


def get_conformite_history_list(db, product_id: int) -> list[dict]:
    from app.models.conformite_history import ConformiteHistory
    rows = (
        db.query(ConformiteHistory)
        .filter(ConformiteHistory.product_id == product_id)
        .order_by(ConformiteHistory.created_at.desc())
        .all()
    )
    return [
        {
            "id": r.id,
            "filename_kelia": r.filename_kelia,
            "filename_contract": r.filename_contract,
            "provider": r.provider,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "kelia_params": r.kelia_params,
            "score_conformite": r.score_conformite,
        }
        for r in rows
    ]


def get_conformite_history_detail(db, history_id: int) -> dict | None:
    import json as _json
    from app.models.conformite_history import ConformiteHistory

    entry = db.query(ConformiteHistory).filter(ConformiteHistory.id == history_id).first()
    if not entry:
        return None
    result = _json.loads(entry.result_json)
    result["history_id"] = entry.id
    result["created_at"] = entry.created_at.isoformat() if entry.created_at else None
    return result


def export_conformite_history_excel(db, history_id: int) -> bytes:
    import io, openpyxl, json as _json
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.models.conformite_history import ConformiteHistory

    entry = db.query(ConformiteHistory).filter(ConformiteHistory.id == history_id).first()
    if not entry:
        raise ValueError(f"Historique {history_id} introuvable")

    result = _json.loads(entry.result_json)
    articles = result.get("articles", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conformité CG"

    header_fill = PatternFill("solid", fgColor="A100FF")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    STATUS_COLORS = {
        "conforme":           "D1FAE5",
        "partiel":            "FEF3C7",
        "non_repris":         "FEE2E2",
        "validation_requise": "EDE9FE",
        "sans_impact":        "F3F4F6",
    }

    headers = ["Article", "Engagement", "Citation CG", "Statut", "Risque type", "Niveau", "Référence FPP", "Explication"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    row_idx = 2
    for art in articles:
        article_label = art.get("article", "")
        for eng in art.get("engagements", []):
            status = eng.get("statut", "")
            row_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
            row_data = [
                article_label,
                eng.get("engagement", ""),
                eng.get("citation_cg", ""),
                status,
                eng.get("risque_type") or "",
                eng.get("risque_niveau", ""),
                eng.get("reference_fpp") or "",
                eng.get("explication", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=(col in (2, 3, 7, 8)))
                if col <= 6:
                    cell.fill = row_fill
            row_idx += 1

    ws_meta = wb.create_sheet("Informations")
    meta = [
        ("Conditions Générales", entry.filename_contract),
        ("FPP", entry.filename_kelia),
        ("Modèle LLM", entry.provider),
        ("Date", entry.created_at.strftime("%d/%m/%Y %H:%M") if entry.created_at else ""),
        ("Score conformité", f"{entry.score_conformite}%"),
        ("Paramètres FPP", entry.kelia_params),
        ("Total engagements", result.get("total_engagements", 0)),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=str(v))

    col_widths = [30, 40, 40, 20, 16, 12, 30, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
