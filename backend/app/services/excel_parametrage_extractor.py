"""
Extracteur dédié pour les fiches de paramétrage BOSS au format structuré.

Détecte automatiquement si un Excel a la structure :
  Domaine BOSS | Menu/écran BOSS | Onglet BOSS | Parametre BOSS | Valeur EXPERIDE | ...

Si oui → lecture directe colonne par colonne, 100 % des règles capturées sans IA.
Si non → retourne None pour laisser l'extraction BMAD classique prendre le relais.
"""
import json
import logging
import re
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Colonnes obligatoires pour considérer la feuille comme une fiche paramétrage
_REQUIRED_HEADERS = {"parametre boss", "valeur"}

# Colonnes optionnelles reconnues (labels normalisés → clé interne)
_HEADER_ALIASES = {
    "domaine boss": "domaine",
    "domaine": "domaine",
    "menu / ecran boss": "menu",
    "menu/ecran boss": "menu",
    "menu": "menu",
    "onglet boss": "onglet",
    "onglet": "onglet",
    "parametre boss": "rule_name",
    "parametre": "rule_name",
    "valeur experide": "rule_value",
    "valeur": "rule_value",
    "description metier": "description",
    "description": "description",
    "source information": "source_doc",
    "source": "source_doc",
    "reference detaillee": "reference",
    "référence détaillée": "reference",
    "confiance": "confiance",
}

# Mapping domaine métier → catégorie 8.x (compatible mapCategory() frontend)
_DOMAINE_TO_CATEGORY = {
    "identification": "8.1 Identification produit",
    "caracteristiques": "8.1 Identification produit",
    "fonds": "8.7 Constitution des droits",
    "cotisations": "8.4 Cotisations et versements",
    "versements": "8.4 Cotisations et versements",
    "frais": "8.6 Frais",
    "arbitrages": "8.6 Frais",
    "dates de valeur": "8.5 Dates de valeur",
    "gestion technique": "8.7 Constitution des droits",
    "gestion financiere": "8.8 Participation aux bénéfices",
    "gestion financière": "8.8 Participation aux bénéfices",
    "gestion rentes": "8.12 Liquidation rente",
    "prestations": "8.10 Rachat",
    "rentes": "8.15 Options rente",
    "transfert": "8.18 Transfert individuel",
    "deces": "8.11 Décès en constitution",
    "décès": "8.11 Décès en constitution",
    "fiscal": "8.19 Fiscalité et taxes",
    "fiscalite": "8.19 Fiscalité et taxes",
    "fiscalité": "8.19 Fiscalité et taxes",
    "information": "8.20 Information assurés",
    "prescription": "8.21 Prescription",
    "si": "8.24 Contraintes SI",
    "parametres techniques": "8.23 Paramètres techniques",
    "paramètres techniques": "8.23 Paramètres techniques",
    "inventaire": "8.22 Inventaire actuariel",
}

# Valeurs NO_VALUE à ignorer
_NO_VALUE_MARKERS = {
    "non renseigné", "non documenté", "à compléter", "nc", "n/a",
    "non applicable", "non mentionné", "non documenté", "",
}

# Confiance textuelle → float
_CONFIANCE_MAP = {
    "certain": 1.0,
    "certain ": 1.0,
    "probable": 0.8,
    "hypothese": 0.5,
    "hypothèse": 0.5,
    "a verifier": 0.3,
    "à vérifier": 0.3,
    "incertain": 0.4,
}


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def _map_category(domaine_raw: str, rule_name: str = "") -> str:
    """
    Map domaine + rule_name to an 8.x category.
    rule_name is used for finer granularity when domaine is broad (e.g. "Rentes").
    """
    d = _norm(domaine_raw)
    r = _norm(rule_name)

    # Fine-grained rules first (rule_name takes precedence for broad domaines)
    if d in ("rentes", "prestations rentes", "rente"):
        if any(k in r for k in ("liquidation", "condition de liquid", "faible rente", "versement unique", "seuil")):
            return "8.12 Liquidation rente"
        if any(k in r for k in ("paiement", "périodicité", "terme", "prorata décès rentier", "certificat")):
            return "8.14 Paiement rente"
        if any(k in r for k in ("revalorisation rente", "revalorisation des rentes")):
            return "8.17 Revalorisation"
        if any(k in r for k in ("fonds collectif", "fonds de service", "fonds de revalorisation")):
            return "8.16 Fonds collectifs"
        if any(k in r for k in ("réversion", "reversion", "annuités", "annuites", "modulable", "dépendance", "dependance", "option")):
            return "8.15 Options rente"
        return "8.12 Liquidation rente"  # default pour "rentes" sans autre indice

    if d in ("prestations",):
        if any(k in r for k in ("rachat", "valeur de rachat", "délai paiement rachat")):
            return "8.10 Rachat"
        if any(k in r for k in ("décès", "deces", "bénéficiaire", "beneficiaire", "capital décès")):
            return "8.11 Décès en constitution"
        if any(k in r for k in ("transfert", "notification", "renonciation")):
            return "8.18 Transfert individuel"
        return "8.10 Rachat"  # default pour "prestations"

    # Standard domaine mapping
    for key, cat in _DOMAINE_TO_CATEGORY.items():
        if key in d:
            return cat
    return domaine_raw or "Autres"


def _parse_unit(value: str) -> tuple[str, Optional[str]]:
    """Split '4,50 %' into ('4,50', '%') or '0,25 %' into ('0,25', '%')."""
    if not value:
        return value, None
    m = re.match(r"^(.*?)\s*(%|EUR|€|ans?|mois|jours?|trimestres?)\s*$", value.strip(), re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2)
    return value.strip(), None


def _confiance_to_float(raw: str) -> float:
    return _CONFIANCE_MAP.get(_norm(raw), 0.85)


def _find_header_row(ws) -> Optional[tuple[int, dict[str, int]]]:
    """
    Scan rows to find the header row.
    Returns (row_index, {col_key: col_index}) or None.
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [_norm(str(c)) if c is not None else "" for c in row]
        mapped: dict[str, int] = {}
        for col_idx, cell in enumerate(cells):
            alias = _HEADER_ALIASES.get(cell)
            if alias and alias not in mapped:
                mapped[alias] = col_idx
        # Must have at least rule_name and rule_value
        if "rule_name" in mapped and "rule_value" in mapped:
            return row_idx, mapped
    return None


def is_parametrage_excel(file_path: str) -> bool:
    """Return True if this Excel looks like a structured BOSS parametrage file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = _find_header_row(ws)
            if result is not None:
                wb.close()
                return True
        wb.close()
    except Exception as e:
        logger.debug(f"is_parametrage_excel check failed for {file_path}: {e}")
    return False


# Valeurs courtes qui ne se suffisent pas à elles-mêmes → description fusionnée dans rule_value
_SHORT_VALUES = {"oui", "non", "oui ", "non ", "vrai", "faux", "présent", "absent",
                 "possible", "interdit", "autorisé", "calcul manuel hors sip",
                 "calcul manuel", "non prévu", "non applicable"}


def _enrich_value(value: str, description: str) -> str:
    """
    Fusionne valeur + description quand la valeur seule est trop courte pour être comprise.
    Ex: "Oui" + "Versements CET investis sur le socle" → "Oui — Versements CET investis sur le socle"
    """
    if not description:
        return value
    if _norm(value) in _SHORT_VALUES or len(value.strip()) <= 4:
        return f"{value} — {description}"
    return value


def _make_unique_name(base: str, qualifier: str, seen: dict[str, int]) -> str:
    """
    Si base déjà vu, ajoute un qualificatif pour différencier.
    Ex: "Réversion" vu 2 fois → "Réversion (CG 8010A)" + "Réversion (NT 0803/2)"
    """
    key = _norm(base)
    if key not in seen:
        seen[key] = 0
        return base
    seen[key] += 1
    q = qualifier.strip()[:30] if qualifier else str(seen[key])
    return f"{base} ({q})"


def extract_from_parametrage_excel(
    file_path: str,
    document_name: str,
) -> list[dict]:
    """
    Extract ALL rules from a structured BOSS parametrage Excel — zero loss.

    Strategy:
    - Reads ALL columns: Parametre, Valeur, Description, Source, Référence, Domaine, Menu, Onglet
    - Description métier enrichit rule_value quand la valeur seule est ambiguë ("Oui", "Calcul manuel"…)
    - Description toujours visible dans source_paragraph
    - Même rule_name avec description/source différent → deux entrées distinctes qualifiées
    - BOSS navigation path (Menu > Onglet) dans subcategory pour traçabilité

    Returns list[dict] compatible with referentiel schema.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — cannot extract parametrage Excel")
        return []

    results: list[dict] = []
    skipped = 0
    seen_names: dict[str, int] = {}  # tracks duplicates across all sheets

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Cannot open {file_path}: {e}")
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_result = _find_header_row(ws)
        if header_result is None:
            logger.debug(f"Sheet '{sheet_name}': no parametrage header found, skipping")
            continue

        header_row_idx, col_map = header_result
        logger.info(f"[excel_parametrage] Sheet '{sheet_name}': header at row {header_row_idx}, cols={col_map}")

        def get(row: tuple, key: str, default: str = "") -> str:
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= header_row_idx:
                continue

            rule_name_raw = get(row, "rule_name")
            rule_value_raw = get(row, "rule_value")
            description = get(row, "description")
            source_doc = get(row, "source_doc")
            reference = get(row, "reference")
            domaine_raw = get(row, "domaine") or sheet_name
            onglet = get(row, "onglet")
            menu = get(row, "menu")
            confiance_raw = get(row, "confiance")

            # --- Filtrage : ignorer lignes vraiment vides ---
            if not rule_name_raw or _norm(rule_name_raw) in _NO_VALUE_MARKERS:
                skipped += 1
                continue
            # Garder si valeur OU description a du contenu — ne jamais jeter une ligne avec description
            has_content = bool(rule_value_raw.strip()) or bool(description.strip())
            if not has_content:
                skipped += 1
                continue

            # --- Valeur enrichie ---
            value_clean, unit = _parse_unit(rule_value_raw)
            if _norm(value_clean) in _NO_VALUE_MARKERS:
                value_clean = "NON TROUVEE"
            # Fusionne description dans value quand la valeur seule est insuffisante
            value_enriched = _enrich_value(value_clean, description)

            # --- Unicité du nom : même paramètre, source différente → deux entrées ---
            # Qualificatif = source document (ex: "NT v1", "CG 8010A") ou index si absent
            qualifier = source_doc.split("/")[0].strip() if source_doc else ""
            rule_name = _make_unique_name(rule_name_raw, qualifier, seen_names)

            # --- Catégorie ---
            category = _map_category(domaine_raw, rule_name_raw)

            # --- Subcategory = chemin BOSS complet ---
            boss_path_parts = [p for p in [menu, onglet] if p]
            subcategory = " > ".join(boss_path_parts) if boss_path_parts else ""

            # --- source_paragraph : Description VISIBLE en premier, puis refs ---
            # C'est la colonne "Source(s)" dans le tableau référentiel — description doit y figurer
            sp_parts = []
            if description:
                sp_parts.append(description)
            if source_doc:
                sp_parts.append(f"Source : {source_doc}")
            if reference:
                sp_parts.append(f"Réf : {reference}")
            source_paragraph = " | ".join(sp_parts)

            confidence = _confiance_to_float(confiance_raw)

            meta = {
                "domaine_boss": domaine_raw,
                "menu_boss": menu,
                "onglet_boss": onglet,
                "boss_path": f"{menu} > {onglet}" if menu and onglet else (menu or onglet or ""),
                "description_metier": description,
                "valeur_brute": rule_value_raw,
                "source_document": source_doc,
                "reference_detaillee": reference,
                "confiance_source": confiance_raw,
                "feuille_excel": sheet_name,
                "extraction_method": "direct_excel_parametrage",
            }

            results.append({
                "category": category,
                "subcategory": subcategory,
                "rule_name": rule_name,
                "rule_value": value_enriched,
                "rule_unit": unit,
                "source_paragraph": source_paragraph,
                "source_page": None,
                "confidence": confidence,
                "comment": json.dumps(meta, ensure_ascii=False),
                "conflict": False,
            })

    wb.close()
    logger.info(
        f"[excel_parametrage] '{document_name}': "
        f"{len(results)} règles extraites, {skipped} lignes ignorées"
    )
    return results
