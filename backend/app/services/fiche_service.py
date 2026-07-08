from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
import openpyxl
import json
import shutil
from datetime import datetime

from app.config import settings
from app.models.fiche import FicheItem
from app.models.product import Product, ProductStatus
from app.models.document import Document
from app.models.referentiel import ReferentielItem
from app.models.ecart import EcartItem
from app.models.version import Version
from app.services import ai_service
from app.services.ai_service import NO_VALUE

import logging

logger = logging.getLogger(__name__)

TEMPLATE_DIR = Path(settings.documents_dir) / "generique"
TEMPLATE_FILENAME = "FPP_KELIA_Template_Model"
TEMPLATE_PATH = TEMPLATE_DIR / f"{TEMPLATE_FILENAME}.xlsx"

# Bullet characters that mark section headers in the template
_BULLET_CHARS = set("●◼►▶⚫○•■▪")


def check_template() -> dict:
    """Check whether the KELIA template file exists and is accessible."""
    exists = TEMPLATE_PATH.exists()
    return {
        "exists": exists,
        "filename": TEMPLATE_FILENAME,
        "path": str(TEMPLATE_PATH),
        "status": "OK" if exists else "INTROUVABLE",
    }


def read_template_structure() -> dict:
    """
    Parse the KELIA template Excel file and return its structure.

    Returns:
        {sheet_name: [{"parameter": str, "section": str,
                        "valeurs_possibles": str|None, "kelia_comment": str|None}, ...]}
    """
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH), data_only=True)
    structure: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fields: list[dict] = []
        current_section = ""

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip first 3 rows (title + subtitle + column headers)
            if row_idx <= 3:
                continue

            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            col_c = row[2] if len(row) > 2 else None
            col_d = row[3] if len(row) > 3 else None

            # Skip empty rows
            if col_a is None:
                continue

            val_a = str(col_a).strip()
            if not val_a:
                continue

            # Detect section header: first char is a bullet char, or both col C and D are None
            first_char = val_a[0] if val_a else ""
            is_bullet = first_char in _BULLET_CHARS

            # Also treat as section header if col B and col C are both empty/None
            # and the text is long or descriptive (i.e. not a short parameter name)
            col_b_empty = col_b is None or str(col_b).strip() == ""
            col_c_empty = col_c is None or str(col_c).strip() == ""
            col_d_empty = col_d is None or str(col_d).strip() == ""

            if is_bullet or (col_b_empty and col_c_empty):
                # This is a section header
                current_section = val_a
                continue

            # It's a parameter field
            fields.append({
                "parameter": val_a,
                "section": current_section,
                "valeurs_possibles": str(col_c).strip() if col_c is not None and str(col_c).strip() else None,
                "kelia_comment": str(col_d).strip() if col_d is not None and str(col_d).strip() else None,
            })

        structure[sheet_name] = fields

    return structure


def generate_fiche(db: Session, product_id: int, complementary_document_ids: list[int] | None = None, referentiel_version: int | None = None) -> list[FicheItem]:
    """
    Generate a Fiche Produit KELIA for the given product.

    Steps:
    1. Validate template exists and product is found
    2. Compute next version number
    3. Load referentiel items for the product
    4. Read template structure
    5. For each sheet, call AI to fill fields from referentiel
    6. Persist FicheItems and return them
    """
    # 1. Check template
    tpl = check_template()
    if not tpl["exists"]:
        raise ValueError(
            f"Modèle Excel introuvable : {tpl['path']}. "
            "Veuillez déposer le fichier FPP_KELIA_Template_Model.xlsx dans le répertoire generique."
        )

    # 2. Load product
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise ValueError(f"Produit {product_id} introuvable.")

    # 3. Compute next version
    max_ver = (
        db.query(sqlfunc.max(FicheItem.version_number))
        .filter(FicheItem.product_id == product_id)
        .scalar()
    )
    next_version = (max_ver or 0) + 1

    # 4. Load referentiel items for the product (specific version or latest)
    from app.models.referentiel import ReferentielItem
    if referentiel_version is not None:
        ref_ver_to_use = referentiel_version
    else:
        ref_ver_to_use = (
            db.query(sqlfunc.max(ReferentielItem.version_number))
            .filter(ReferentielItem.product_id == product_id)
            .scalar()
        )
    if ref_ver_to_use is None:
        ref_items_db = []
    else:
        ref_items_db = (
            db.query(ReferentielItem)
            .filter(
                ReferentielItem.product_id == product_id,
                ReferentielItem.version_number == ref_ver_to_use,
            )
            .all()
        )

    # Build referentiel context list
    # Pre-load documents for source filenames
    doc_ids_in_ref = set()
    for ri in ref_items_db:
        if ri.source_document_id:
            doc_ids_in_ref.add(ri.source_document_id)
    docs_by_id: dict[int, Document] = {}
    if doc_ids_in_ref:
        for doc in db.query(Document).filter(Document.id.in_(list(doc_ids_in_ref))).all():
            docs_by_id[doc.id] = doc

    ref_context: list[dict] = []
    for ri in ref_items_db:
        src_doc_filename = ""
        if ri.source_document_id and ri.source_document_id in docs_by_id:
            src_doc_filename = docs_by_id[ri.source_document_id].original_filename
        ref_context.append({
            "category": ri.category or "",
            "rule_name": ri.rule_name,
            "rule_value": ri.rule_value or "",
            "source_doc": src_doc_filename,
            "confidence": ri.ai_confidence or 0.0,
            "conflict": bool(ri.conflict),
            "source_document_id": ri.source_document_id,
            "source_document_ids": ri.source_document_ids,
            "source_paragraph": ri.source_paragraph or "",
        })

    # Build lookup dicts keyed by rule_name (lower) for traceability
    # ref_lookup_multi keeps ALL items per rule (multi-source support)
    ref_lookup_multi: dict[str, list[dict]] = {}
    for rc in ref_context:
        key = rc["rule_name"].strip().lower()
        if key not in ref_lookup_multi:
            ref_lookup_multi[key] = []
        ref_lookup_multi[key].append(rc)
    # Single-entry fallback for code paths that still need one item
    ref_lookup: dict[str, dict] = {k: v[0] for k, v in ref_lookup_multi.items()}

    # 5. Extract complementary document items (if any) — these take priority over referentiel
    cr_items: list[dict] = []
    cr_lookup: dict[str, dict] = {}
    comp_docs_by_id: dict[int, Document] = {}

    if complementary_document_ids:
        for comp_doc in db.query(Document).filter(Document.id.in_(complementary_document_ids)).all():
            comp_docs_by_id[comp_doc.id] = comp_doc
            text = comp_doc.extracted_text or ""
            if not text:
                logger.warning(f"[fiche_service] Complementary doc {comp_doc.id} ({comp_doc.original_filename}) has no extracted text")
                continue
            extracted = ai_service.extract_cr_atelier(text, product.boss_number)
            for item in extracted:
                enriched = {**item, "document_id": comp_doc.id, "filename": comp_doc.original_filename}
                cr_items.append(enriched)
                key = (item.get("rule_name") or "").strip().lower()
                if key and key not in cr_lookup:
                    cr_lookup[key] = enriched
        logger.info(f"[fiche_service] {len(cr_items)} items extraits depuis {len(complementary_document_ids)} doc(s) complémentaire(s)")

    # 6. Read template structure
    try:
        template_structure = read_template_structure()
    except Exception as e:
        raise ValueError(f"Impossible de lire le modèle Excel : {e}")

    # 7. Fill each sheet via AI (pre-mapping done per sheet)
    all_items: list[FicheItem] = []
    all_mapping_tables: list[dict] = []

    for sheet_name, fields in template_structure.items():
        if not fields:
            continue

        logger.info(f"[fiche_service] Remplissage feuille '{sheet_name}' ({len(fields)} champs)")

        # Pre-mapping pass for this sheet only (~55 fields → fits in 4096 tokens)
        sheet_mapping: list[dict] = []
        if ref_context:
            try:
                sheet_mapping = ai_service.pre_mapping_pass(
                    all_fields=fields,
                    referentiel_items=ref_context,
                )
                all_mapping_tables.extend(sheet_mapping)
                logger.info(f"[fiche_service] Pré-mapping '{sheet_name}': {len(sheet_mapping)} correspondances")
            except Exception as e:
                logger.error(f"[fiche_service] Erreur pré-mapping '{sheet_name}': {e}")

        try:
            ai_results = ai_service.fill_fiche_sheet(
                sheet_name=sheet_name,
                fields=fields,
                referentiel_items=ref_context,
                product_number=product.boss_number,
                cr_items=cr_items if cr_items else None,
                mapping_table=sheet_mapping if sheet_mapping else None,
            )
        except Exception as e:
            logger.error(f"[fiche_service] Erreur AI pour feuille '{sheet_name}': {e}")
            ai_results = []

        # Build a lookup dict from AI results by parameter name
        ai_lookup: dict[str, dict] = {}
        for ai_item in ai_results:
            param_key = (ai_item.get("parameter") or "").strip().lower()
            if param_key:
                ai_lookup[param_key] = ai_item

        # Create a FicheItem for each template field
        for field in fields:
            param_key = field["parameter"].strip().lower()
            ai_item = ai_lookup.get(param_key)

            if ai_item:
                value = ai_item.get("value") or NO_VALUE
                source_paragraph = ai_item.get("source_paragraph")
                source_citation = ai_item.get("source_citation")
                ai_confidence = ai_item.get("confidence")
                ai_comment = ai_item.get("comment")
                conflict = bool(ai_item.get("conflict", False))
                is_cr_override = bool(ai_item.get("cr_override", False))
            else:
                value = NO_VALUE
                source_paragraph = None
                source_citation = None
                ai_confidence = None
                ai_comment = None
                conflict = False
                is_cr_override = False

            source_document_id = None
            source_document_ids_json = None

            if is_cr_override and ai_item:
                # Source is the CR document — look up by cr_rule_matched
                cr_rule_matched = ai_item.get("cr_rule_matched")
                matched_cr = cr_lookup.get(cr_rule_matched.strip().lower()) if cr_rule_matched else None
                if matched_cr:
                    cr_doc_id = matched_cr.get("document_id")
                    cr_filename = matched_cr.get("filename", "")
                    cr_para = matched_cr.get("source_paragraph", "")
                    source_document_id = cr_doc_id
                    source_document_ids_json = json.dumps([cr_doc_id]) if cr_doc_id else None
                    source_paragraph = f"CR Atelier — {cr_filename}" + (f" — {cr_para}" if cr_para else "")
                conflict = True  # CR override always implies conflict (divergence from referentiel)

            elif ai_item and not is_cr_override:
                # Source is from the referentiel — collect ALL matching sources (multi-doc support)
                rule_name_ref = ai_item.get("rule_name_ref")
                if rule_name_ref:
                    lookup_key = rule_name_ref.strip().lower()
                elif source_paragraph:
                    lookup_key = source_paragraph.replace("Règle référentiel:", "").strip().lower()
                else:
                    lookup_key = None

                all_matches = ref_lookup_multi.get(lookup_key, []) if lookup_key else []

                if all_matches:
                    # Collect unique sources (deduplicate by doc_id)
                    seen_doc_ids: set = set()
                    sources_list: list[dict] = []
                    all_doc_ids_found: list[int] = []

                    for matched in all_matches:
                        doc_id = matched.get("source_document_id")
                        stored_ids_raw = matched.get("source_document_ids")
                        # Resolve doc IDs for this match
                        match_doc_ids: list[int] = []
                        if stored_ids_raw:
                            try:
                                match_doc_ids = json.loads(stored_ids_raw)
                            except Exception:
                                pass
                        if not match_doc_ids and doc_id:
                            match_doc_ids = [doc_id]

                        for did in match_doc_ids:
                            if did not in seen_doc_ids:
                                seen_doc_ids.add(did)
                                all_doc_ids_found.append(did)

                        # Resolve filename
                        fname = ""
                        if doc_id and doc_id in docs_by_id:
                            fname = docs_by_id[doc_id].original_filename
                        elif doc_id:
                            fname = matched.get("source_doc", "")

                        para = matched.get("source_paragraph") or ""

                        # Deduplicate by (fname, para) pair
                        entry_key = (fname, para[:80])
                        if not any(
                            (s.get("doc") == fname and s.get("text", "")[:80] == para[:80])
                            for s in sources_list
                        ):
                            sources_list.append({"doc": fname, "text": para, "doc_id": doc_id})

                        if matched.get("conflict"):
                            conflict = True

                    # Set source_document_id to first, source_document_ids to all
                    if all_doc_ids_found:
                        source_document_id = all_doc_ids_found[0]
                        source_document_ids_json = json.dumps(all_doc_ids_found)

                    # Encode sources: JSON array if multiple, plain string if single
                    valid_sources = [s for s in sources_list if s.get("doc") or s.get("text")]
                    if len(valid_sources) > 1:
                        source_paragraph = json.dumps(valid_sources, ensure_ascii=False)
                    elif len(valid_sources) == 1:
                        s = valid_sources[0]
                        if s.get("doc") and s.get("text"):
                            source_paragraph = f"{s['doc']} — {s['text']}"
                        elif s.get("doc"):
                            source_paragraph = s["doc"]
                        else:
                            source_paragraph = s.get("text", "")

            fiche_item = FicheItem(
                product_id=product_id,
                version_number=next_version,
                sheet=sheet_name,
                section=field.get("section") or "",
                parameter=field["parameter"],
                valeurs_possibles=field.get("valeurs_possibles"),
                kelia_comment=field.get("kelia_comment"),
                value=value,
                source_document_id=source_document_id,
                source_document_ids=source_document_ids_json,
                source_paragraph=source_paragraph,
                source_citation=source_citation,
                ai_confidence=ai_confidence,
                ai_comment=ai_comment,
                conflict=conflict,
                cr_override=is_cr_override,
            )
            db.add(fiche_item)
            all_items.append(fiche_item)

    # 7. Validate at least some real data was extracted
    real_values = [i for i in all_items if i.value and i.value != NO_VALUE]
    if len(real_values) == 0:
        db.rollback()
        raise ValueError(
            "Aucune donnée extraite depuis le référentiel. "
            "Vérifiez que le référentiel produit a bien été généré avant la fiche."
        )

    # ── ÉCARTS : règles détectées dans le référentiel mais non mappées à un champ FPP ──────────
    # Collect all parameter names used in the fiche (case-insensitive)
    fiche_params_used: set[str] = set()
    for fi in all_items:
        if fi.value and fi.value != NO_VALUE:
            fiche_params_used.add((fi.parameter or "").strip().lower())
            # Also track the source_paragraph to match referentiel items
            if fi.source_paragraph:
                fiche_params_used.add(fi.source_paragraph.strip().lower()[:100])

    # Build set of referentiel rule_names that were mapped to FPP fields
    mapped_rule_names: set[str] = set()
    for mapping_entry in all_mapping_tables:
        rn = (mapping_entry.get("rule_name") or "").strip().lower()
        if rn:
            mapped_rule_names.add(rn)

    # Pre-load document names for écarts
    ecart_doc_names: dict[int, str] = {}
    for doc in db.query(Document).filter(Document.product_id == product_id).all():
        ecart_doc_names[doc.id] = doc.original_filename

    # Find unmapped referentiel items → Écarts
    ecart_items: list[EcartItem] = []
    for rc in ref_context:
        rn_lower = (rc.get("rule_name") or "").strip().lower()
        # Skip noisy / empty items
        rv = (rc.get("rule_value") or "").strip()
        if not rv or rv.lower() in ("non trouvee dans ce document", "a verifier", "non trouvée"):
            continue
        # If this rule_name was NOT used in any fiche field mapping → ÉCART
        if rn_lower and rn_lower not in mapped_rule_names:
            src_doc_id = rc.get("source_document_id")
            src_doc_ids_raw = rc.get("source_document_ids")
            if not src_doc_id and src_doc_ids_raw:
                try:
                    ids = json.loads(src_doc_ids_raw)
                    src_doc_id = ids[0] if ids else None
                except Exception:
                    pass
            ecart = EcartItem(
                product_id=product_id,
                fiche_version_number=next_version,
                rule_name=rc.get("rule_name") or "",
                rule_value=rv,
                category=rc.get("category"),
                source_document_id=src_doc_id,
                source_document_name=ecart_doc_names.get(src_doc_id, "") if src_doc_id else rc.get("source_doc", ""),
                source_page=None,
                source_section=None,
                source_paragraph=(rc.get("source_paragraph") or "")[:1000],
                ecart_type="UNMAPPED_RULE",
                ai_confidence=rc.get("confidence"),
            )
            db.add(ecart)
            ecart_items.append(ecart)

    logger.info(
        f"[fiche_service] Écarts V{next_version} produit {product_id}: "
        f"{len(ref_context)} règles référentiel → {len(ecart_items)} non mappées FPP"
    )

    # ── CR PRIORITÉ ABSOLUE : forcer cr_override sur tous les champs CR ────────────────────────
    # CR items always override referentiel — already handled above per field, but ensure consistency
    for fi in all_items:
        if fi.cr_override:
            fi.conflict = True   # CR override always signals a divergence visible

    # 8. Update product status
    product.status_fiche = ProductStatus.GENERATED

    # 9. Create Version record for this fiche generation (tracks all doc sources)
    # Total doc IDs: those from the referentiel version + complementary docs
    ref_doc_ids: list[int] = []
    if ref_ver_to_use is not None:
        ref_ver_obj = db.query(Version).filter(
            Version.product_id == product_id,
            Version.artifact_type == "Referentiel",
            Version.version_number == ref_ver_to_use,
        ).first()
        if ref_ver_obj and ref_ver_obj.document_ids:
            try:
                ref_doc_ids = json.loads(ref_ver_obj.document_ids)
            except Exception:
                pass

    all_doc_ids = list(dict.fromkeys(ref_doc_ids + (complementary_document_ids or [])))
    fiche_version = Version(
        product_id=product_id,
        artifact_type="Fiche",
        version_number=next_version,
        version_label=f"V{next_version}",
        document_ids=json.dumps(all_doc_ids),
        snapshot={
            "referentiel_version": ref_ver_to_use,
            "complementary_document_ids": complementary_document_ids or [],
            "ref_document_ids": ref_doc_ids,
            "mapping_table": all_mapping_tables,
        },
    )
    db.add(fiche_version)

    db.commit()
    logger.info(
        f"[fiche_service] Fiche V{next_version} générée pour produit {product_id}: "
        f"{len(all_items)} champs, {len(real_values)} renseignés, "
        f"{len(all_doc_ids)} docs tracés"
    )
    return all_items


def export_fiche_excel(db: Session, product_id: int) -> str:
    """
    Export the latest Fiche Produit to an Excel file based on the KELIA template.

    Returns the path of the generated file.
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise ValueError(f"Produit {product_id} introuvable.")

    # Get latest version
    max_ver = (
        db.query(sqlfunc.max(FicheItem.version_number))
        .filter(FicheItem.product_id == product_id)
        .scalar()
    )
    if max_ver is None:
        raise ValueError("Aucune fiche générée pour ce produit.")

    items = (
        db.query(FicheItem)
        .filter(FicheItem.product_id == product_id, FicheItem.version_number == max_ver)
        .all()
    )
    if not items:
        raise ValueError("Aucun champ trouvé pour la fiche de ce produit.")

    # Build a lookup dict: (sheet_name, parameter) -> FicheItem
    item_lookup: dict[tuple, FicheItem] = {}
    for fi in items:
        item_lookup[(fi.sheet, fi.parameter)] = fi

    # Copy template to exports dir
    exports_dir = Path(settings.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_filename = f"FicheKELIA_{product.boss_number}_V{max_ver}_{timestamp}.xlsx"
    export_path = exports_dir / export_filename

    shutil.copy2(str(TEMPLATE_PATH), str(export_path))

    # Open the copied workbook and fill in values
    wb = openpyxl.load_workbook(str(export_path))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_idx = 0
        for row in ws.iter_rows():
            row_idx += 1
            if row_idx <= 3:
                continue
            col_a_cell = row[0] if len(row) > 0 else None
            if col_a_cell is None or col_a_cell.value is None:
                continue

            param_name = str(col_a_cell.value).strip()
            if not param_name:
                continue

            key = (sheet_name, param_name)
            fiche_item = item_lookup.get(key)
            if fiche_item and fiche_item.value and fiche_item.value != NO_VALUE:
                # Write value to col B
                col_b_cell = row[1] if len(row) > 1 else None
                if col_b_cell is not None:
                    col_b_cell.value = fiche_item.value

                # Write source_paragraph to col E (index 4)
                if len(row) > 4:
                    row[4].value = fiche_item.source_paragraph or ""
                else:
                    # Append a cell in col E if the row doesn't have it
                    ws.cell(row=row_idx, column=5, value=fiche_item.source_paragraph or "")

    wb.save(str(export_path))
    logger.info(f"[fiche_service] Export Excel: {export_path}")
    return str(export_path)
