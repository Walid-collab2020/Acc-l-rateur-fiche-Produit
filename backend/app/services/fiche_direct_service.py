"""
Fiche Produit — Génération directe depuis les documents sources.
Nouveau moteur : 4 appels LLM parallèles (un par onglet FPP), prompt expert Actuaire/MOA.
"""
from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
import json
import shutil
import logging
import re
import openpyxl
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

import anthropic
from openai import OpenAI

from app.config import settings
from app.models.fiche_direct import FicheDirectItem, FicheExtraInfo
from app.models.document import Document
from app.models.product import Product, ProductStatus
from app.models.version import Version
from app.services.ai_service import get_active_provider

logger = logging.getLogger(__name__)

NO_VALUE = "Information manquante"
TEMPLATE_PATH = Path(settings.storage_dir) / "documents" / "generique" / "FPP_KELIA_Template_Model.xlsx"

SHEETS_TO_PROCESS = [
    "Produit Technique",
    "Tarif de Rente",
    "Garanties et Prestations",
    "Mode de Gestion",
]

# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """Tu es un expert combinant les rôles suivants :
- Actuaire Assurance Vie, Épargne et Retraite
- Expert MOA Assurance
- Expert Paramétrage KELIA
- Analyste Fonctionnel
- Expert de la réglementation française en assurance (Art.83, PERCO, PER, PERP, L.132-5, L.144-1, L.144-2, A.160-2...)

Ton objectif est de remplir un onglet de Fiche de Paramétrage Produit (FPP) KELIA à partir de documents sources fournis.
Le résultat doit être justifiable, traçable, auditable et exploitable par les équipes MOA, Actuariat, Paramétrage et Recette.

PRINCIPES ABSOLUS :
1. Tu ne génères jamais une valeur sans justification documentaire.
2. Tu ne tranches jamais seul une contradiction entre documents — tu la signales.
3. Tu n'inventes aucune information. Si une donnée est absente : "Information manquante".
4. Les documents dont le nom contient "CR" sont prioritaires en cas de contradiction, mais toutes les valeurs contradictoires restent affichées.
5. Tu respectes strictement les valeurs KELIA autorisées quand elles sont fournies."""

USER_PROMPT_TEMPLATE = """## DOCUMENTS SOURCE

{documents_text}

---

## TEMPLATE FPP — ONGLET À REMPLIR : {sheet_name}

{sheet_template}

---

## INSTRUCTIONS DE GÉNÉRATION

### PHASE 1 — Compréhension globale (obligatoire avant tout remplissage)

Avant de traiter les paramètres, construis une vision d'ensemble du produit :
- Nature du produit, cadre fiscal, compagnie portante
- Garanties principales et optionnelles identifiées
- Règles métier clés (revalorisation, liquidation, rachat, réversion...)
- Contradictions détectées entre documents
- Informations manquantes critiques

### PHASE 2 — Remplissage des paramètres

Pour chaque paramètre du template :
1. Cherche l'information dans l'ensemble des documents.
2. Respecte strictement les valeurs KELIA possibles si fournies.
3. Tiens compte des commentaires et points de vigilance.
4. Si l'information est absente : valeur = "Information manquante", explique pourquoi.

RÈGLE ABSOLUE — EXTRACTION BOSS / FICHE PARAMÉTRAGE :
Dans l'extraction BOSS, la colonne "Valeur [PRODUIT]" contient la valeur paramétrée en production dans KELIA.
La "Description metier" est un commentaire contextuel — elle peut être incomplète ou erronée.
TOUJOURS utiliser la "Valeur [PRODUIT]" de l'extraction BOSS comme valeur de production, même si la description dit autre chose.
Exemples :
- Si "Valeur [PRODUIT]: 3 %" et "Description: Aucun frais" → la valeur est 3 %, la description est fausse.
- Si "Valeur [PRODUIT]: 0,50 % annuel" → la valeur est 0,50%, quel que soit le contexte.

RÈGLE ABSOLUE — CONTRADICTIONS ENTRE DOCUMENTS :
Quand au moins un document donne une valeur DIFFÉRENTE des autres pour le MÊME paramètre :
- Génère DEUX entrées séparées dans "fields" avec le MÊME "parameter"
- Première entrée : valeur majoritaire (BOSS / CG), source_document = nom exact
- Deuxième entrée : valeur divergente (ex : Note Technique), source_document = nom exact
- Les deux entrées ont "contradiction": {{"detected": true, "details": "DocA: valeurX / DocB: valeurY"}}
- Ne jamais choisir une seule valeur quand des documents sont en désaccord
- L'extraction BOSS ou la fiche de paramétrage existante représente le paramétrage en production : capturer TOUJOURS sa valeur
- RÈGLE CRITIQUE : si deux documents s'accordent (ex : BOSS=0,5% et CG=0,5%) mais qu'un troisième donne une valeur différente (ex : Note Technique=1%), c'est TOUJOURS une contradiction — générer 2 entrées. NE PAS ignorer le document divergent sous prétexte que les deux autres s'accordent.
- Exemple FGE : BOSS=0,5%, CG=0,5%, NT=1% → entrée 1 "0,5%" (BOSS ou CG), entrée 2 "1%" (Note Technique), les deux avec contradiction=true

ÉTAPE OBLIGATOIRE AVANT LE JSON — TABLEAU DE COMPARAISON :
Avant de générer le JSON, pour CHAQUE paramètre du template contenant une valeur numérique (%, EUR, taux, durée), écris un tableau de comparaison au format :
  PARAM: <nom paramètre> | BOSS: <valeur ou N/A> | CG: <valeur ou N/A> | NT: <valeur ou N/A> | → <Valide ou CONTRADICTION>

Règles du tableau :
- Lis chaque document SÉPARÉMENT pour ce paramètre avant de remplir la colonne
- Si un document ne mentionne pas ce paramètre : N/A
- Si les valeurs sont identiques entre tous les documents : → Valide
- Si AU MOINS UN document donne une valeur différente : → CONTRADICTION (générer 2 entrées JSON)
- Ce tableau est OBLIGATOIRE — sans lui, tu risques de manquer des contradictions

Exemple attendu :
  PARAM: Frais sur encours (gestion) % | BOSS: 0,50% | CG: 0,50% | NT: 1% | → CONTRADICTION
  PARAM: Frais sur versement % | BOSS: 4,50% | CG: 4,50% | NT: 4% | → CONTRADICTION
  PARAM: TMG | BOSS: N/A | CG: 0% | NT: 0% | → Valide

SCORES DE CONFIANCE :
- 100% : valeur explicitement présente, aucune interprétation, aucune contradiction
- 90% : valeur explicite, léger rapprochement entre éléments du même document
- 75% : valeur déduite de plusieurs éléments convergents
- 50% : hypothèse métier cohérente, non explicitement documentée
- 25% : fortement incertain, validation humaine obligatoire
- 0% : aucune information exploitable

CONTRAINTES DE CONCISION (obligatoires pour éviter la troncature) :
- justification : 1 phrase max (20 mots)
- source_extract : 1 phrase extraite mot pour mot du document (30 mots max)

RÈGLES ABSOLUES SUR LES VALEURS :
- Les "(ex: ...)" dans le template sont des exemples de format, PAS des valeurs par défaut. N'utilise un exemple comme valeur que s'il est confirmé par les documents sources. Si non confirmé : "Information manquante".
- Raisonne avant chaque valeur : est-ce que cette valeur est explicitement ou implicitement documentée ?
- Modes de règlement : n'indique que ce qui est cité dans les documents (Chèque, Virement, Espèces). Liste tous les modes cités. N'en invente pas.
- Règle d'effet et délais : uniquement si documentés, sinon "Information manquante".
- Durée du TMG : durée d'application du taux minimum garanti (en années ou par génération de strate), sans lien avec les rentes viagères.
{garanties_instruction}
### PHASE 3 — Paramètres orphelins (obligatoire)

Après le remplissage, identifie TOUS les paramètres présents dans les documents mais ABSENTS du template FPP.
Recherche notamment :
- Frais non listés (frais sur encours fixes, frais d'arrérage, frais de transfert, frais CET...)
- Méthodes de calcul spécifiques (intérêts techniques continus vs discrets, interpolation vs âge arrondi...)
- Règles de gestion absentes (ordre de sortie strates FIFO, délais légaux, certificats de vie annuels...)
- Seuils réglementaires (faible rente, plafonds, tranches de cotisation TA/TB/TC/TD...)
- Paramètres de migration (codes SI historiques, dates migration, gestionnaire opérationnel...)
- Tout autre paramètre métier utile au paramétrage ou à la recette KELIA

---

## FORMAT DE RÉPONSE OBLIGATOIRE

Réponds UNIQUEMENT en JSON valide, sans texte avant ni après le JSON.
IMPORTANT : sois CONCIS — chaque champ texte doit faire au maximum 20 mots.

{{
  "sheet": "{sheet_name}",
  "product_understanding": "1 phrase résumant le produit",
  "fields": [
    {{
      "parameter": "Nom exact du paramètre",
      "guarantee_block": null,
      "value": "Valeur ou Information manquante",
      "confidence_pct": 90,
      "justification": "1 phrase max",
      "source_document": "Nom du doc",
      "source_page": 3,
      "source_extract": "Citation exacte 15 mots max",
      "hypotheses": null,
      "contradiction": {{
        "detected": false,
        "details": null
      }}
    }},
    {{
      "parameter": "EXEMPLE CONTRADICTION — même paramètre, doc différent",
      "guarantee_block": null,
      "value": "Valeur du second document (OBLIGATOIRE si contradiction)",
      "confidence_pct": 75,
      "justification": "Valeur issue du second document",
      "source_document": "Nom du second doc",
      "source_page": 5,
      "source_extract": "Citation exacte du second doc",
      "hypotheses": null,
      "contradiction": {{
        "detected": true,
        "details": "PremierDoc: valeurA / SecondDoc: valeurB"
      }}
    }}
  ],
  "extra_information": [
    {{
      "parameter": "Nom",
      "value": "Valeur",
      "source_document": "Doc",
      "source_page": 1,
      "source_extract": "Citation courte",
      "comment": "1 phrase",
      "recommendation": "Ajouter à la FPP"
    }}
  ],
  "open_points": [
    {{
      "code": "E-001",
      "description": "1 phrase",
      "impact": "1 phrase",
      "action": "1 phrase"
    }}
  ]
}}"""

GARANTIES_INSTRUCTION = """
### INSTRUCTIONS SPÉCIALES — Garanties et Prestations

Identifie TOUTES les garanties à paramétrer dans KELIA :
- Garanties principales (ex: Rente viagère différée)
- Garanties optionnelles (réversion, annuités garanties, dépendance, rente modulable...)
- Garanties en cas de décès
- Prestations accessoires

Pour chaque garantie identifiée, duplique le bloc de paramètres du template et renseigne le champ "guarantee_block" avec le nom de la garantie.
Chaque bloc doit être traité de manière INDÉPENDANTE et COMPLÈTE.
"""


# ---------------------------------------------------------------------------
# Détection des types de documents manquants
# ---------------------------------------------------------------------------

REQUIRED_DOC_TYPES = [
    {
        "doc_type": "Note Technique Actuarielle",
        "severity": "critique",
        "message": "Contient les taux techniques, tables de mortalite, formules de rente — parametres actuariels critiques.",
        "recommendation": "Sans ce document, les parametres actuariels (TMG, calcul rente, tables) seront absents.",
    },
    {
        "doc_type": "Conditions Generales",
        "severity": "critique",
        "message": "Contient les droits assures, clauses de rachat/transfert/deces/reversion.",
        "recommendation": "Sans ce document, les regles contractuelles et legales seront absentes.",
    },
]

IMPORTANT_DOC_TYPES = [
    {
        "doc_type": "Extraction BOSS",
        "severity": "important",
        "message": "Contient les parametres SI source (codes BOSS/KELIA, donnees de gestion).",
        "recommendation": "Sans ce document, les codes SI et les parametres de parametrage seront manquants.",
    },
    {
        "doc_type": "Fiche Produit",
        "severity": "important",
        "message": "Contient les parametres de la fiche produit existante.",
        "recommendation": "Ce document permet de verifier la coherence avec le parametrage existant.",
    },
    {
        "doc_type": "Compte-rendu Atelier",
        "severity": "recommande",
        "message": "Contient les decisions et corrections validees en atelier.",
        "recommendation": "Les decisions CR priment sur les autres documents.",
    },
]

_DOC_TYPE_ALIASES = {
    "conditions generales": "Conditions Generales",
    "conditions générales": "Conditions Generales",
    "note technique actuarielle": "Note Technique Actuarielle",
    "note technique": "Note Technique Actuarielle",
    "extraction boss": "Extraction BOSS",
    "fiche produit": "Fiche Produit",
    "parametrage kelia": "Fiche Produit",
    "paramétrage kelia": "Fiche Produit",
    "compte-rendu atelier": "Compte-rendu Atelier",
    "compte rendu atelier": "Compte-rendu Atelier",
}


def check_missing_document_types(documents: list[Document]) -> list[dict]:
    present_types: set[str] = set()
    for doc in documents:
        cat = (doc.category or "").lower().strip()
        normalized = _DOC_TYPE_ALIASES.get(cat, doc.category or "")
        present_types.add(normalized)

    warnings: list[dict] = []
    for req in REQUIRED_DOC_TYPES:
        if req["doc_type"] not in present_types:
            warnings.append({
                "type": "missing_document_type",
                "severity": req["severity"],
                "doc_type": req["doc_type"],
                "message": req["message"],
                "recommendation": req["recommendation"],
            })
    for imp in IMPORTANT_DOC_TYPES:
        if imp["doc_type"] not in present_types:
            warnings.append({
                "type": "missing_document_type",
                "severity": imp["severity"],
                "doc_type": imp["doc_type"],
                "message": imp["message"],
                "recommendation": imp["recommendation"],
            })
    return warnings


# ---------------------------------------------------------------------------
# Lecture du template FPP
# ---------------------------------------------------------------------------

def _read_template_sheets() -> dict[str, str]:
    """Lit le template Excel et retourne {sheet_name: texte_formaté_pour_prompt}."""
    if not TEMPLATE_PATH.exists():
        raise ValueError(f"Template FPP introuvable : {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(str(TEMPLATE_PATH), data_only=True)
    result: dict[str, str] = {}

    for sheet_name in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            result[sheet_name] = ""
            continue

        ws = wb[sheet_name]
        lines: list[str] = []

        for row in ws.iter_rows(min_row=5, values_only=True):
            param = str(row[0] or "").strip()
            valeurs = str(row[2] or "").strip() if len(row) > 2 else ""
            comment = str(row[3] or "").strip() if len(row) > 3 else ""

            if not param or param == "None":
                continue

            if param.startswith("█"):
                lines.append(f"\n### {param}")
            else:
                line = f"- PARAMÈTRE: {param}"
                if valeurs and valeurs != "None":
                    line += f"\n  VALEURS KELIA: {valeurs}"
                if comment and comment != "None":
                    line += f"\n  VIGILANCE: {comment}"
                lines.append(line)

        result[sheet_name] = "\n".join(lines)

    return result


def _read_template_item_list() -> dict[str, list[dict]]:
    """Retourne {sheet_name: [{parameter, valeurs_possibles, kelia_comment, section}]}."""
    if not TEMPLATE_PATH.exists():
        raise ValueError(f"Template FPP introuvable : {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(str(TEMPLATE_PATH), data_only=True)
    result: dict[str, list[dict]] = {}
    current_section = ""

    for sheet_name in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            result[sheet_name] = []
            continue

        ws = wb[sheet_name]
        fields: list[dict] = []

        for row in ws.iter_rows(min_row=5, values_only=True):
            param = str(row[0] or "").strip()
            valeurs = str(row[2] or "").strip() if len(row) > 2 else ""
            comment = str(row[3] or "").strip() if len(row) > 3 else ""

            if not param or param == "None":
                continue

            if param.startswith("█"):
                current_section = param.lstrip("█ ").strip()
                continue

            fields.append({
                "parameter": param,
                "section": current_section,
                "valeurs_possibles": valeurs if valeurs != "None" else "",
                "kelia_comment": comment if comment != "None" else "",
            })

        result[sheet_name] = fields
        current_section = ""

    return result


# ---------------------------------------------------------------------------
# Construction du contexte documentaire
# ---------------------------------------------------------------------------

def _build_documents_context(documents: list[Document]) -> str:
    """Construit le texte combiné de tous les documents pour le prompt."""
    parts: list[str] = []
    for doc in documents:
        text = (doc.extracted_text or "").strip()
        if not text:
            continue
        is_cr = "CR" in (doc.original_filename or "").upper()
        priority_marker = " ⭐ PRIORITAIRE (Compte-rendu)" if is_cr else ""
        header = f"=== DOCUMENT : {doc.original_filename}{priority_marker} ===\nCatégorie : {doc.category or 'Non classifié'}\nPages : {doc.page_count or 'N/A'}\n\n"
        parts.append(header + text)

    return "\n\n" + ("=" * 80) + "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Appel LLM
# ---------------------------------------------------------------------------

def _call_llm(system: str, user: str, max_tokens: int = 16000) -> str:
    """Appel LLM vers le provider actif avec system + user prompt."""
    provider = get_active_provider()

    if provider == "anthropic":
        client = anthropic.Anthropic(api_key=settings.anthropic_api_key or None)
        resp = client.messages.create(
            model=settings.anthropic_model,
            max_tokens=max_tokens,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        return resp.content[0].text.strip()
    else:
        client = OpenAI(api_key=settings.openai_api_key or None)
        resp = client.chat.completions.create(
            model=settings.openai_model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            max_tokens=max_tokens,
            temperature=0,
            response_format={"type": "json_object"},
        )
        return resp.choices[0].message.content.strip()


def _sanitize_json(raw: str) -> str:
    """Extrait le JSON d'une réponse LLM (supprime texte avant/après les accolades)."""
    raw = raw.strip()
    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]
    # Supprime les caractères de contrôle illégaux
    raw = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', raw)
    return raw


CHUNK_SIZE = 25  # Champs max par appel LLM pour eviter la troncature JSON


def _fields_to_template_text(fields: list[dict]) -> str:
    """Convertit une liste de champs template en texte formaté pour le prompt."""
    lines: list[str] = []
    current_section = ""
    for f in fields:
        sec = f.get("section", "")
        if sec and sec != current_section:
            current_section = sec
            lines.append(f"\n### {sec}")
        line = f"- PARAMETRE: {f['parameter']}"
        if f.get("valeurs_possibles"):
            line += f"\n  VALEURS KELIA: {f['valeurs_possibles']}"
        if f.get("kelia_comment"):
            line += f"\n  VIGILANCE: {f['kelia_comment']}"
        lines.append(line)
    return "\n".join(lines)


def _call_llm_chunk(sheet_name: str, chunk_fields: list[dict], docs_context: str,
                    product_context: str = "", chunk_info: str = "") -> dict:
    """Appel LLM pour un sous-ensemble de champs d'un onglet."""
    is_garanties = sheet_name == "Garanties et Prestations"
    garanties_instr = GARANTIES_INSTRUCTION if is_garanties else ""
    chunk_template = _fields_to_template_text(chunk_fields)

    context_note = ""
    if product_context:
        context_note = f"\n\nCONTEXTE PRODUIT : {product_context}\n"
    if chunk_info:
        context_note += f"\nATTENTION : {chunk_info}\n"

    user_prompt = USER_PROMPT_TEMPLATE.format(
        documents_text=docs_context,
        sheet_name=sheet_name,
        sheet_template=chunk_template + context_note,
        garanties_instruction=garanties_instr,
    )

    raw = ""
    try:
        raw = _call_llm(SYSTEM_PROMPT, user_prompt, max_tokens=16000)
        raw = _sanitize_json(raw)
        return json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error(f"[fiche_direct] JSON invalide chunk '{sheet_name}': {e}\nRaw[:300]: {raw[:300]}")
        return {"sheet": sheet_name, "fields": [], "extra_information": [], "open_points": [], "error": str(e)}
    except Exception as e:
        logger.error(f"[fiche_direct] Erreur LLM chunk '{sheet_name}': {e}", exc_info=True)
        return {"sheet": sheet_name, "fields": [], "extra_information": [], "open_points": [], "error": str(e)}


def _call_llm_for_sheet(sheet_name: str, template_text: str, docs_context: str,
                        template_fields: list[dict] | None = None) -> dict:
    """
    Lance les appels LLM pour un onglet FPP.
    Pour les onglets > CHUNK_SIZE champs, decoupe en plusieurs appels sequentiels.
    """
    fields = template_fields or []

    # Garanties : appel unique (blocs detectes dynamiquement)
    if sheet_name == "Garanties et Prestations" or len(fields) <= CHUNK_SIZE:
        logger.info(f"[fiche_direct] Appel unique '{sheet_name}' ({len(fields)} champs)...")
        return _call_llm_chunk(sheet_name, fields, docs_context)

    # Decoupage en chunks pour les grands onglets
    chunks = [fields[i:i + CHUNK_SIZE] for i in range(0, len(fields), CHUNK_SIZE)]
    logger.info(f"[fiche_direct] '{sheet_name}' : {len(fields)} champs -> {len(chunks)} chunks")

    all_fields: list[dict] = []
    all_extra: list[dict] = []
    all_open_points: list[dict] = []
    product_understanding = ""

    for idx, chunk in enumerate(chunks):
        chunk_info = f"Partie {idx+1}/{len(chunks)} — remplir UNIQUEMENT ces {len(chunk)} parametres."
        result = _call_llm_chunk(
            sheet_name, chunk, docs_context,
            product_context=product_understanding,
            chunk_info=chunk_info,
        )
        if not product_understanding and result.get("product_understanding"):
            product_understanding = result["product_understanding"]
        all_fields.extend(result.get("fields", []))
        if idx == len(chunks) - 1:
            all_extra.extend(result.get("extra_information", []))
            all_open_points.extend(result.get("open_points", []))

    logger.info(f"[fiche_direct] '{sheet_name}' total : {len(all_fields)} champs apres {len(chunks)} chunks")
    return {
        "sheet": sheet_name,
        "product_understanding": product_understanding,
        "fields": all_fields,
        "extra_information": all_extra,
        "open_points": all_open_points,
    }


# ---------------------------------------------------------------------------
# Calcul du statut métier
# ---------------------------------------------------------------------------

def _compute_status(value: str, confidence_pct: int | None, contradiction: bool) -> str:
    if not value or value == NO_VALUE or value == "Information manquante":
        return "Information manquante"
    if contradiction:
        return "Sources contradictoires"
    if confidence_pct is None:
        return "A verifier"
    if confidence_pct >= 90:
        return "Valide"
    if confidence_pct >= 70:
        return "A verifier"
    if confidence_pct >= 50:
        return "Ambigu"
    return "A verifier"


# ---------------------------------------------------------------------------
# Stockage des résultats
# ---------------------------------------------------------------------------

def _store_sheet_results(
    db: Session,
    product_id: int,
    version_number: int,
    sheet_result: dict,
    template_fields: list[dict],
) -> list[FicheDirectItem]:
    """
    Stocke les résultats d'un onglet LLM dans la DB.
    Pour les onglets standards : on croise avec le template pour ne pas rater de champs.
    Pour Garanties et Prestations : on stocke les blocs tels que retournés (duplication incluse).
    """
    sheet_name = sheet_result.get("sheet", "")
    llm_fields: list[dict] = sheet_result.get("fields") or []
    is_garanties = sheet_name == "Garanties et Prestations"

    def _norm(s: str) -> str:
        """Normalise un nom de paramètre : strip espaces et astérisques finaux."""
        return s.strip().rstrip("*").strip()

    # Index LLM par (param_normalisé, guarantee_block) → liste (supporte contradictions = 2 entrées même param)
    llm_lookup: dict[tuple[str, str | None], list[dict]] = {}
    llm_lookup_norm: dict[tuple[str, str | None], list[dict]] = {}
    for f in llm_fields:
        param = (f.get("parameter") or "").strip()
        block = f.get("guarantee_block") or None
        if param:
            key_exact = (param, block)
            key_norm = (_norm(param), block)
            llm_lookup.setdefault(key_exact, []).append(f)
            llm_lookup_norm.setdefault(key_norm, []).append(f)

    items: list[FicheDirectItem] = []

    if is_garanties:
        # Pour Garanties et Prestations : on prend tous les blocs retournés par le LLM
        for llm_field in llm_fields:
            item = _build_fdi(product_id, version_number, sheet_name, llm_field, template_fields)
            db.add(item)
            items.append(item)
    else:
        # Pour les autres onglets : on itère sur le template pour garantir tous les champs
        # Si le LLM a renvoyé plusieurs entrées pour le même paramètre (contradiction), on crée N lignes
        for tpl_field in template_fields:
            param = tpl_field["parameter"]
            matched: list[dict] = (
                llm_lookup.get((param, None))
                or llm_lookup.get((param.strip(), None))
                or llm_lookup_norm.get((_norm(param), None))
                or []
            )

            if not matched:
                # Champ non retourné par le LLM — ligne vide
                llm_field = {
                    "parameter": param,
                    "value": NO_VALUE,
                    "confidence_pct": 0,
                    "justification": "Non retourné par le LLM",
                }
                item = _build_fdi(product_id, version_number, sheet_name, llm_field, template_fields, tpl_field)
                db.add(item)
                items.append(item)
            else:
                # 1 entrée (cas normal) ou N entrées (contradiction entre documents)
                batch: list[FicheDirectItem] = []
                for llm_field in matched:
                    llm_field = dict(llm_field)
                    llm_field["parameter"] = param  # nom exact du template
                    item = _build_fdi(product_id, version_number, sheet_name, llm_field, template_fields, tpl_field)
                    db.add(item)
                    items.append(item)
                    batch.append(item)

                # Auto-split : LLM a renvoyé 1 seule entrée avec contradiction=true
                # et les détails contiennent "DocA: valX / DocB: valY" → créer la 2e ligne
                if len(batch) == 1 and batch[0].conflict:
                    second = _split_contradiction_item(
                        batch[0], product_id, version_number, sheet_name, template_fields, tpl_field
                    )
                    if second is not None:
                        db.add(second)
                        items.append(second)

    return items


def _split_contradiction_item(
    item: FicheDirectItem,
    product_id: int,
    version_number: int,
    sheet_name: str,
    template_fields: list[dict],
    tpl_field: dict,
) -> "FicheDirectItem | None":
    """
    Quand le LLM renvoie 1 seule entrée contradiction avec details='DocA: valX / DocB: valY',
    construit synthétiquement la 2e ligne (DocB / valY) pour l'affichage.
    """
    detail = item.contradiction_detail or ""
    if " / " not in detail:
        return None
    parts = detail.split(" / ", 1)
    if len(parts) != 2:
        return None
    second_part = parts[1]
    colon_idx = second_part.find(": ")
    if colon_idx < 0:
        return None
    doc2 = second_part[:colon_idx].strip()
    val2 = second_part[colon_idx + 2:].strip()
    if not val2 or not doc2:
        return None
    llm_field2: dict = {
        "parameter": item.parameter,
        "value": val2,
        "confidence_pct": item.confidence_pct,
        "justification": f"Valeur issue de {doc2}",
        "source_document": doc2,
        "source_page": None,
        "source_extract": None,
        "contradiction": {"detected": True, "details": detail},
    }
    return _build_fdi(product_id, version_number, sheet_name, llm_field2, template_fields, tpl_field)


def _build_fdi(
    product_id: int,
    version_number: int,
    sheet_name: str,
    llm_field: dict,
    template_fields: list[dict],
    tpl_field: dict | None = None,
) -> FicheDirectItem:
    """Construit un FicheDirectItem depuis les données LLM + template."""
    param = (llm_field.get("parameter") or "").strip()
    guarantee_block = llm_field.get("guarantee_block") or None

    # Trouver les métadonnées du template pour ce paramètre
    if tpl_field is None:
        tpl_field = next((t for t in template_fields if t["parameter"] == param), {})

    raw_value = llm_field.get("value")
    value = (str(raw_value) if raw_value is not None else NO_VALUE).strip()
    if not value or value.lower() in ("null", "none"):
        value = NO_VALUE

    confidence_pct = llm_field.get("confidence_pct")
    if isinstance(confidence_pct, (int, float)) and not isinstance(confidence_pct, bool):
        confidence_pct = max(0, min(100, int(confidence_pct)))
    else:
        confidence_pct = None

    contradiction_info = llm_field.get("contradiction")
    if isinstance(contradiction_info, dict):
        contradiction = bool(contradiction_info.get("detected", False))
        contradiction_detail = contradiction_info.get("details") or None
    elif isinstance(contradiction_info, bool):
        contradiction = contradiction_info
        contradiction_detail = None
    else:
        contradiction = False
        contradiction_detail = None

    status = _compute_status(value, confidence_pct, contradiction)

    # Section : guarantee_block pour Garanties, sinon section du template
    section = guarantee_block or tpl_field.get("section") or ""

    # Safely parse source_page as int
    raw_page = llm_field.get("source_page")
    source_page: int | None = None
    if raw_page is not None:
        try:
            source_page = int(float(str(raw_page)))
        except (ValueError, TypeError):
            source_page = None

    return FicheDirectItem(
        product_id=product_id,
        version_number=version_number,
        sheet=sheet_name,
        section=section,
        parameter=param,
        valeurs_possibles=tpl_field.get("valeurs_possibles") or "",
        kelia_comment=tpl_field.get("kelia_comment") or "",
        value=value,
        status=status,
        source_document_id=None,
        source_paragraph=llm_field.get("source_document") or None,
        source_citation=llm_field.get("source_extract") or None,
        source_page=source_page,
        sources_json=None,
        ai_confidence=(confidence_pct / 100.0) if confidence_pct is not None else None,
        ai_comment=llm_field.get("justification") or None,
        conflict=contradiction,
        # Nouveaux champs traçabilité
        confidence_pct=confidence_pct,
        justification=llm_field.get("justification") or None,
        reasoning=llm_field.get("reasoning") or None,
        source_extract=llm_field.get("source_extract") or None,
        hypotheses=llm_field.get("hypotheses") or None,
        contradiction_detail=contradiction_detail,
    )


def _parse_pct_value(s: str) -> float | None:
    """Extrait un float depuis '0,50 %', '1%', '3 %', etc."""
    if not s:
        return None
    m = re.search(r'(\d+(?:[,\.]\d+)?)\s*%', s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(',', '.'))
    except ValueError:
        return None


def _find_pct_near_keywords(text: str, keywords: list[str]) -> tuple[float | None, str | None]:
    """
    Cherche une valeur % dans le texte à proximité de plusieurs mots-clés.
    Retourne (float_pct, str_pct) ou (None, None).
    """
    lines = text.split('\n')
    kw_lower = [k.lower() for k in keywords]
    for i, line in enumerate(lines):
        ll = line.lower()
        if sum(1 for k in kw_lower if k in ll) >= 2:
            # Cherche un % dans la ligne ou les 2 lignes suivantes
            context = '\n'.join(lines[i:min(len(lines), i + 3)])
            m = re.search(r'(\d+(?:[,\.]\d+)?)\s*%', context)
            if m:
                raw = m.group(0).strip()
                val = _parse_pct_value(raw)
                return val, raw
    return None, None


def _post_process_contradictions(
    db: Session,
    product_id: int,
    version_number: int,
    documents: list,
) -> None:
    """
    Post-traitement après génération LLM : détecte les contradictions de valeurs %
    manquées par le LLM en comparant les textes de tous les documents.

    Algorithme :
    - Pour chaque item "Valide" avec une valeur % (frais, taux…)
    - Pour chaque autre document : chercher les mots-clés du paramètre + extraire %
    - Si % différent trouvé → convertir item en contradiction + créer 2e ligne
    """
    items = db.query(FicheDirectItem).filter(
        FicheDirectItem.product_id == product_id,
        FicheDirectItem.version_number == version_number,
        FicheDirectItem.status == "Valide",
        FicheDirectItem.conflict == False,  # noqa: E712
    ).all()

    if not items:
        return

    # Mots d'arrêt pour extraction de mots-clés
    STOP = {'sur', 'de', 'du', 'des', 'le', 'la', 'les', 'et', 'en', 'par', 'taux',
            'frais', 'pct', 'gestion', 'nets', 'net', 'ou'}

    for item in items:
        item_pct = _parse_pct_value(item.value or "")
        if item_pct is None:
            continue

        # Mots-clés du paramètre (> 3 lettres, hors stop words)
        words = re.findall(r'\b[a-zA-Zàâéèêëîïôùûüç]+\b', (item.parameter or "").lower())
        keywords = [w for w in words if len(w) > 3 and w not in STOP]
        if len(keywords) < 1:
            continue

        source_doc_name = (item.source_paragraph or "").lower()

        for doc in documents:
            if not doc.extracted_text:
                continue
            # Ignorer le document source de cet item
            doc_name = (doc.original_filename or "").lower()
            if doc_name and source_doc_name and (
                doc_name in source_doc_name or source_doc_name in doc_name
            ):
                continue

            found_pct, found_str = _find_pct_near_keywords(doc.extracted_text, keywords)
            if found_pct is None:
                continue
            if abs(found_pct - item_pct) < 0.001:
                continue  # Même valeur → pas de contradiction

            # Contradiction détectée !
            detail = f"{item.source_paragraph or 'Source'}: {item.value} / {doc.original_filename}: {found_str}"
            logger.info(
                f"[post_process] Contradiction détectée — '{item.parameter}': "
                f"{item.value} ({item.source_paragraph}) vs {found_str} ({doc.original_filename})"
            )

            # Mettre à jour l'item existant
            item.conflict = True
            item.contradiction_detail = detail
            item.status = "Sources contradictoires"

            # Créer la 2e ligne
            item2 = FicheDirectItem(
                product_id=item.product_id,
                version_number=item.version_number,
                sheet=item.sheet,
                section=item.section,
                parameter=item.parameter,
                valeurs_possibles=item.valeurs_possibles or "",
                kelia_comment=item.kelia_comment or "",
                value=found_str,
                status="Sources contradictoires",
                source_document_id=None,
                source_paragraph=doc.original_filename,
                source_citation=None,
                source_page=None,
                sources_json=None,
                ai_confidence=0.75,
                ai_comment=f"Valeur issue de {doc.original_filename}",
                conflict=True,
                confidence_pct=75,
                justification=f"Valeur issue de {doc.original_filename}",
                reasoning=None,
                source_extract=None,
                hypotheses=None,
                contradiction_detail=detail,
            )
            db.add(item2)
            break  # Un seul doc contradictoire suffit pour créer la 2e ligne


def _store_extra_info(
    db: Session,
    product_id: int,
    version_number: int,
    sheet_results: list[dict],
) -> list[FicheExtraInfo]:
    """Stocke les paramètres orphelins et points ouverts dans FicheExtraInfo."""
    extras: list[FicheExtraInfo] = []
    seen: set[str] = set()

    for sheet_result in sheet_results:
        for item in (sheet_result.get("extra_information") or []):
            key = f"{item.get('parameter', '')}|{item.get('value', '')}"
            if key in seen:
                continue
            seen.add(key)
            raw_p = item.get("source_page")
            ei_page: int | None = None
            if raw_p is not None:
                try:
                    ei_page = int(float(str(raw_p)))
                except (ValueError, TypeError):
                    ei_page = None
            ei = FicheExtraInfo(
                product_id=product_id,
                version_number=version_number,
                parameter=item.get("parameter") or "",
                value=item.get("value") or "",
                source_document=item.get("source_document") or "",
                source_page=ei_page,
                source_extract=item.get("source_extract") or "",
                comment=item.get("comment") or "",
                recommendation=item.get("recommendation") or "A arbitrer",
                is_open_point=False,
            )
            db.add(ei)
            extras.append(ei)

        for op in (sheet_result.get("open_points") or []):
            ei = FicheExtraInfo(
                product_id=product_id,
                version_number=version_number,
                parameter=op.get("description") or "",
                value="",
                source_document="",
                comment=op.get("impact") or "",
                recommendation="Point ouvert",
                is_open_point=True,
                open_point_code=op.get("code") or "",
                open_point_impact=op.get("impact") or "",
                open_point_action=op.get("action") or "",
            )
            db.add(ei)
            extras.append(ei)

    return extras


# ---------------------------------------------------------------------------
# Orchestrateur principal
# ---------------------------------------------------------------------------

def analyze_and_fill_fpp(
    db: Session,
    product_id: int,
    document_ids: list[int],
    sheets_to_process: list[str] | None = None,
) -> tuple[list[FicheDirectItem], list[dict]]:
    """
    Moteur FPP : appels LLM séquentiels (un par onglet) pour fiabilité maximale.
    sheets_to_process : liste des onglets à traiter (None = les 4 par défaut).
    Retourne (items, warnings).
    """
    if not TEMPLATE_PATH.exists():
        raise ValueError(f"Template FPP introuvable : {TEMPLATE_PATH}")

    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise ValueError(f"Produit {product_id} introuvable.")

    documents = db.query(Document).filter(Document.id.in_(document_ids)).all()
    if not documents:
        raise ValueError("Aucun document valide trouvé.")

    # Vérification documents manquants
    warnings = check_missing_document_types(documents)

    # Lecture template + documents
    template_sheets_text = _read_template_sheets()
    template_items = _read_template_item_list()
    docs_context = _build_documents_context(documents)

    if not docs_context.strip():
        raise ValueError(
            "Aucun texte extrait dans les documents sélectionnés. "
            "Vérifiez que les documents ont été traités (bouton Ré-extraire)."
        )

    # Onglets à traiter (filtre si l'utilisateur en a choisi un seul)
    active_sheets = [s for s in SHEETS_TO_PROCESS if sheets_to_process is None or s in sheets_to_process]

    # Version suivante
    max_ver = (
        db.query(sqlfunc.max(FicheDirectItem.version_number))
        .filter(FicheDirectItem.product_id == product_id)
        .scalar()
    )
    next_version = (max_ver or 0) + 1

    # Appels LLM séquentiels (fiabilité > parallélisme)
    sheet_results: dict[str, dict] = {}
    for sheet_name in active_sheets:
        logger.info(f"[fiche_direct] Traitement onglet '{sheet_name}'...")
        try:
            sheet_results[sheet_name] = _call_llm_for_sheet(
                sheet_name,
                template_sheets_text.get(sheet_name, ""),
                docs_context,
                template_items.get(sheet_name, []),
            )
        except Exception as e:
            logger.error(f"[fiche_direct] Echec onglet '{sheet_name}': {e}")
            sheet_results[sheet_name] = {
                "sheet": sheet_name,
                "fields": [],
                "extra_information": [],
                "open_points": [],
                "error": str(e),
            }

    # Stockage en DB
    all_items: list[FicheDirectItem] = []
    for sheet_name in active_sheets:
        result = sheet_results.get(sheet_name, {})
        try:
            items = _store_sheet_results(
                db, product_id, next_version, result, template_items.get(sheet_name, [])
            )
            all_items.extend(items)
        except Exception as e:
            logger.error(f"[fiche_direct] Erreur stockage '{sheet_name}': {type(e).__name__}: {e}", exc_info=True)
            raise ValueError(f"Erreur lors du stockage de l'onglet '{sheet_name}': {type(e).__name__}: {e}")

    # Stockage des paramètres orphelins et points ouverts
    try:
        _store_extra_info(db, product_id, next_version, list(sheet_results.values()))
    except Exception as e:
        logger.warning(f"[fiche_direct] Erreur stockage extra_info (non-bloquant): {e}", exc_info=True)

    # Post-traitement : détection contradictions manquées par le LLM (comparaison % inter-documents)
    try:
        _post_process_contradictions(db, product_id, next_version, documents)
    except Exception as e:
        logger.warning(f"[fiche_direct] Erreur post-process contradictions (non-bloquant): {e}", exc_info=True)

    filled = sum(1 for i in all_items if i.value and i.value not in (NO_VALUE, "Information manquante"))
    if filled == 0:
        db.rollback()
        raise ValueError(
            "Aucune valeur renseignée. "
            "Vérifiez que les documents contiennent du texte extrait."
        )

    # Snapshot de version
    product_understanding = " | ".join(
        str(r.get("product_understanding", ""))
        for r in sheet_results.values()
        if r.get("product_understanding") and isinstance(r.get("product_understanding"), (str, int, float))
    )[:2000]

    product.status_fiche = ProductStatus.GENERATED
    fiche_version = Version(
        product_id=product_id,
        artifact_type="FicheDirect",
        version_number=next_version,
        version_label=f"V{next_version}",
        document_ids=json.dumps(document_ids),
        snapshot={
            "document_ids": document_ids,
            "warnings": warnings,
            "product_understanding": product_understanding,
            "sheet_errors": {
                name: r.get("error")
                for name, r in sheet_results.items()
                if r.get("error")
            },
        },
    )
    db.add(fiche_version)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"[fiche_direct] Erreur commit final: {type(e).__name__}: {e}", exc_info=True)
        raise ValueError(f"Erreur lors de la sauvegarde en base : {type(e).__name__}: {e}")

    logger.info(
        f"[fiche_direct] V{next_version} produit {product_id}: "
        f"{len(all_items)} champs, {filled} renseignés, "
        f"{len(warnings)} warning(s)"
    )
    return all_items, warnings


# ---------------------------------------------------------------------------
# Rétrocompatibilité (anciens endpoints /extract + /fill)
# ---------------------------------------------------------------------------

def extract_rules_for_direct(
    db: Session, product_id: int, document_ids: list[int]
) -> tuple[int, list[dict]]:
    """Délègue vers le nouveau moteur. Retourne (version, warnings)."""
    items, warnings = analyze_and_fill_fpp(db, product_id, document_ids)
    max_ver = (
        db.query(sqlfunc.max(FicheDirectItem.version_number))
        .filter(FicheDirectItem.product_id == product_id)
        .scalar()
    ) or 1
    return max_ver, warnings


def fill_fiche_from_extracted(
    db: Session, product_id: int, extraction_version: int
) -> tuple[list[FicheDirectItem], list[dict]]:
    """Retourne les items déjà stockés à cette version (le remplissage s'est fait dans extract)."""
    items = (
        db.query(FicheDirectItem)
        .filter(
            FicheDirectItem.product_id == product_id,
            FicheDirectItem.version_number == extraction_version,
        )
        .all()
    )
    if not items:
        raise ValueError(f"Version V{extraction_version} introuvable pour le produit {product_id}.")
    return items, []


def generate_fiche_direct(
    db: Session, product_id: int, document_ids: list[int]
) -> tuple[list[FicheDirectItem], list[dict]]:
    """Wrapper rétrocompat."""
    return analyze_and_fill_fpp(db, product_id, document_ids)


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def export_fiche_direct_excel(db: Session, product_id: int) -> str:
    max_ver = (
        db.query(sqlfunc.max(FicheDirectItem.version_number))
        .filter(FicheDirectItem.product_id == product_id)
        .scalar()
    )
    if max_ver is None:
        raise ValueError("Aucune Fiche Produit générée pour ce produit.")

    product = db.query(Product).filter(Product.id == product_id).first()
    items = (
        db.query(FicheDirectItem)
        .filter(FicheDirectItem.product_id == product_id, FicheDirectItem.version_number == max_ver)
        .all()
    )
    extra_infos = (
        db.query(FicheExtraInfo)
        .filter(FicheExtraInfo.product_id == product_id, FicheExtraInfo.version_number == max_ver)
        .all()
    )

    item_lookup: dict[tuple[str, str], FicheDirectItem] = {}
    for fi in items:
        key = (fi.sheet, fi.parameter)
        if key not in item_lookup:
            item_lookup[key] = fi

    exports_dir = Path(settings.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prod_label = product.boss_number if product else str(product_id)
    filename = f"FPP_{prod_label}_V{max_ver}_{timestamp}.xlsx"
    export_path = exports_dir / filename
    shutil.copy2(str(TEMPLATE_PATH), str(export_path))

    wb = openpyxl.load_workbook(str(export_path))

    # Remplissage des onglets existants
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEETS_TO_PROCESS:
            continue
        ws = wb[sheet_name]

        # Ajouter les en-têtes des colonnes de traçabilité (colonnes F à L)
        header_row = None
        for row in ws.iter_rows(min_row=4, max_row=4):
            header_row = row
            break

        if header_row and len(header_row) > 4:
            trace_headers = [
                "Confiance (%)", "Justification", "Raisonnement",
                "Document source", "Page", "Extrait source", "Hypothèses",
            ]
            for i, h in enumerate(trace_headers):
                col_idx = 5 + i  # colonnes F à L (index 5 à 11, 0-based)
                cell = ws.cell(row=4, column=col_idx + 1)
                cell.value = h

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            if row_idx <= 3:
                continue
            col_a = row[0] if row else None
            if not col_a or not col_a.value:
                continue
            param_name = str(col_a.value).strip()
            if param_name.startswith("█") or param_name == "PARAMÈTRE":
                continue

            fi = item_lookup.get((sheet_name, param_name))
            if fi and fi.value and fi.value not in (NO_VALUE, "Information manquante"):
                if len(row) > 1:
                    row[1].value = fi.value
                # Traçabilité
                trace_values = [
                    fi.confidence_pct,
                    fi.justification,
                    fi.reasoning,
                    fi.source_paragraph,  # nom du doc source
                    fi.source_page,
                    fi.source_extract,
                    fi.hypotheses,
                ]
                for i, val in enumerate(trace_values):
                    cell = ws.cell(row=row_idx, column=6 + i)
                    cell.value = val

    # Onglet Informations Complémentaires
    if extra_infos:
        ic_ws = wb.create_sheet("Informations Complémentaires")
        headers = [
            "Paramètre", "Valeur", "Document source", "Page",
            "Extrait", "Commentaire", "Recommandation", "Décision",
            "Point ouvert", "Code", "Impact", "Action recommandée",
        ]
        ic_ws.append(headers)
        for ei in extra_infos:
            ic_ws.append([
                ei.parameter, ei.value, ei.source_document, ei.source_page,
                ei.source_extract, ei.comment, ei.recommendation, ei.user_decision or "",
                "Oui" if ei.is_open_point else "Non",
                ei.open_point_code or "",
                ei.open_point_impact or "",
                ei.open_point_action or "",
            ])

    # Scrubbing demo : supprimer termes sensibles de toutes les cellules
    _scrub_workbook(wb)

    wb.save(str(export_path))
    return str(export_path)


_SCRUB_TERMS = [
    ("Expéride", "Produit"),
    ("EXPERIDE", "PRODUIT"),
    ("Experide", "Produit"),
    ("21 rue Laffitte, 75317 Paris Cedex 09", ""),
    ("21 rue Laffitte 75317 Paris Cedex 09", ""),
    ("21 rue Laffitte", ""),
    ("75317 Paris Cedex 09", ""),
    ("Paris Cedex 09", ""),
]


def _scrub_workbook(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    v = cell.value
                    for old, new in _SCRUB_TERMS:
                        v = v.replace(old, new)
                    v = v.strip().strip(",").strip()
                    cell.value = v if v else None
