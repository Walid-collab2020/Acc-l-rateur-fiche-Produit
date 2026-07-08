import os
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.config import settings


COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "section_bg": "BDD7EE",
    "conforme_bg": "C6EFCE",
    "ecart_bg": "FFC7CE",
    "manquant_bg": "FFEB9C",
    "supplementaire_bg": "E2EFDA",
    "non_controlable_bg": "D9D9D9",
}

STATUS_COLORS = {
    "Conforme": "C6EFCE",
    "Écart": "FFC7CE",
    "Manquant": "FFEB9C",
    "Supplémentaire": "E2EFDA",
    "Non contrôlable": "D9D9D9",
}


def _header_style(cell, bg_color: str = COLORS["header_bg"], fg_color: str = COLORS["header_fg"]):
    cell.font = Font(name="Calibri", bold=True, color=fg_color, size=10)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _data_style(cell, bg_color: str = "FFFFFF"):
    cell.font = Font(name="Calibri", size=9)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def export_referentiel(product_number: str, items: list[dict], output_path: Optional[str] = None) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Référentiel {product_number}"

    headers = [
        "Catégorie", "Sous-catégorie", "Règle", "Valeur", "Unité",
        "Source", "Page", "Extrait", "Confiance IA", "Commentaire IA"
    ]
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        _header_style(ws.cell(1, col))

    categories = {}
    for item in items:
        cat = item.get("category", "Autres")
        categories.setdefault(cat, []).append(item)

    row = 2
    for cat, cat_items in categories.items():
        for item in cat_items:
            ws.cell(row, 1, cat)
            ws.cell(row, 2, item.get("subcategory", ""))
            ws.cell(row, 3, item.get("rule_name", ""))
            ws.cell(row, 4, item.get("rule_value", ""))
            ws.cell(row, 5, item.get("rule_unit", ""))
            ws.cell(row, 6, item.get("source_document", ""))
            ws.cell(row, 7, item.get("source_page", ""))
            ws.cell(row, 8, item.get("source_paragraph", ""))
            conf = item.get("ai_confidence") or item.get("confidence", 0)
            ws.cell(row, 9, f"{conf:.0%}" if conf else "")
            ws.cell(row, 10, item.get("ai_comment") or item.get("comment", ""))
            bg = "F2F2F2" if row % 2 == 0 else "FFFFFF"
            for col in range(1, 11):
                _data_style(ws.cell(row, col), bg)
            row += 1

    col_widths = [20, 20, 40, 30, 10, 30, 8, 50, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path(settings.exports_dir) / f"Referentiel_{product_number}_{ts}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_recette(product_number: str, recette_data: dict, anomalies: list[dict], output_path: Optional[str] = None) -> str:
    wb = openpyxl.Workbook()

    # Sheet 1: Synthèse
    ws_syn = wb.active
    ws_syn.title = "Synthèse"

    ws_syn["A1"] = f"RAPPORT DE RECETTE - Produit {product_number}"
    ws_syn["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_syn["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    summary_data = [
        ["", ""],
        ["Indicateur", "Valeur"],
        ["Total contrôles", recette_data.get("total_controls", 0)],
        ["Conformes", recette_data.get("conformes", 0)],
        ["Écarts", recette_data.get("ecarts", 0)],
        ["Régressions", recette_data.get("regressions", 0)],
        ["Anomalies critiques", recette_data.get("anomalies_critiques", 0)],
        ["Anomalies majeures", recette_data.get("anomalies_majeures", 0)],
        ["Anomalies mineures", recette_data.get("anomalies_mineures", 0)],
        ["Taux de conformité", f"{recette_data.get('taux_conformite', 0):.1%}"],
    ]
    for r, row_data in enumerate(summary_data, 3):
        for c, val in enumerate(row_data, 1):
            ws_syn.cell(r, c, val)
            if r == 4:
                _header_style(ws_syn.cell(r, c), COLORS["header_bg"])
            elif c == 1 and r > 4:
                _data_style(ws_syn.cell(r, c), "F2F2F2")
            else:
                _data_style(ws_syn.cell(r, c))

    ws_syn.column_dimensions["A"].width = 30
    ws_syn.column_dimensions["B"].width = 20

    # Sheet 2: Journal des écarts
    ws_ecarts = wb.create_sheet("Journal des écarts")
    ecart_headers = [
        "Produit", "Version", "Module", "Règle", "Valeur attendue",
        "Valeur obtenue", "Statut", "Criticité", "Source", "Page",
        "Commentaire IA", "Commentaire analyste", "Statut anomalie"
    ]
    ws_ecarts.append(ecart_headers)
    for col, h in enumerate(ecart_headers, 1):
        _header_style(ws_ecarts.cell(1, col))

    for row_num, anom in enumerate(anomalies, 2):
        ws_ecarts.cell(row_num, 1, product_number)
        ws_ecarts.cell(row_num, 2, anom.get("version", ""))
        ws_ecarts.cell(row_num, 3, anom.get("module", ""))
        ws_ecarts.cell(row_num, 4, anom.get("rule_name", ""))
        ws_ecarts.cell(row_num, 5, anom.get("expected_value", ""))
        ws_ecarts.cell(row_num, 6, anom.get("obtained_value", ""))
        status = anom.get("status", "")
        ws_ecarts.cell(row_num, 7, status)
        ws_ecarts.cell(row_num, 8, anom.get("criticite", ""))
        ws_ecarts.cell(row_num, 9, anom.get("source_document", ""))
        ws_ecarts.cell(row_num, 10, anom.get("source_page", ""))
        ws_ecarts.cell(row_num, 11, anom.get("ai_comment", ""))
        ws_ecarts.cell(row_num, 12, anom.get("analyst_comment", ""))
        ws_ecarts.cell(row_num, 13, anom.get("anomalie_status", ""))
        bg = STATUS_COLORS.get(status, "FFFFFF")
        for col in range(1, 14):
            _data_style(ws_ecarts.cell(row_num, col), bg)

    for i, w in enumerate([12, 10, 20, 40, 30, 30, 15, 12, 30, 8, 40, 40, 15], 1):
        ws_ecarts.column_dimensions[get_column_letter(i)].width = w
    ws_ecarts.freeze_panes = "A2"

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path(settings.exports_dir) / f"Recette_{product_number}_{ts}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_fiche_produit(product_number: str, items: list[dict], output_path: Optional[str] = None) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fiche Produit {product_number}"

    headers = ["Section", "Sous-section", "Paramètre KELIA", "Valeur", "Justification", "Source", "Page", "Confiance", "Commentaire", "Statut validation"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(1, col))

    for row_num, item in enumerate(items, 2):
        ws.cell(row_num, 1, item.get("section", ""))
        ws.cell(row_num, 2, item.get("subsection", ""))
        ws.cell(row_num, 3, item.get("rule_name", ""))
        ws.cell(row_num, 4, item.get("rule_value", ""))
        ws.cell(row_num, 5, item.get("justification", ""))
        ws.cell(row_num, 6, item.get("source_document", ""))
        ws.cell(row_num, 7, item.get("source_page", ""))
        conf = item.get("ai_confidence") or item.get("confidence", 0)
        ws.cell(row_num, 8, f"{conf:.0%}" if conf else "")
        ws.cell(row_num, 9, item.get("ai_comment") or item.get("comment", ""))
        ws.cell(row_num, 10, item.get("validated", "En attente"))
        bg = "F2F2F2" if row_num % 2 == 0 else "FFFFFF"
        for col in range(1, 11):
            _data_style(ws.cell(row_num, col), bg)

    for i, w in enumerate([25, 25, 40, 30, 50, 30, 8, 12, 40, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path(settings.exports_dir) / f"FicheProduit_{product_number}_{ts}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
