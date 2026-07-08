"""
Service de recette paramétrage.

1. get_fpp_versions         → liste les versions FPP disponibles pour un produit
2. compare_fpp_versions     → non-régression entre deux versions FPP (pure DB)
3. parse_parametrage_file   → parse un fichier KELIA uploadé (Excel/CSV)
4. _llm_match_and_compare   → matching sémantique LLM + comparaison valeurs
5. compare_parametrage_vs_fpp → matrice KELIA ↔ FPP (1 ligne par ligne KELIA)
"""
from __future__ import annotations

import io
import csv
import json
import re
import unicodedata
from typing import Any

from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc

from app.models.fiche_direct import FicheDirectItem
from app.models.version import Version

NO_VAL = "Information manquante"
_EMPTY = {NO_VAL, "Aucune regle mentionnee dans les documents analyses", "", None}

_STOP_WORDS = {"pour", "avec", "sans", "dans", "leur", "code", "type", "date",
               "mode", "base", "calcul", "liste", "valeur", "montant", "taux",
               "sur", "les", "des", "par", "une", "est", "son", "ses"}

# ── Normalisation ─────────────────────────────────────────────────────────────

def _norm(s: str | None) -> str:
    """Lowercase, sans accents, sans ponctuation — pour matching fuzzy."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip().rstrip("*").strip()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _word_sim(a: str, b: str) -> float:
    """Score Jaccard sur mots importants (>= 4 chars, sans stopwords, dé-pluralisé)."""
    def words(s: str) -> set[str]:
        result = set()
        for w in s.split():
            if w in _STOP_WORDS or len(w) < 4:
                continue
            if w.endswith("s") and len(w) > 4:
                w = w[:-1]
            result.add(w)
        return result

    wa, wb = words(a), words(b)
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / max(len(wa), len(wb))


def _is_empty(val: str | None) -> bool:
    return (val or "").strip() in _EMPTY or not (val or "").strip()


def _numeric_equiv(a: str | None, b: str | None) -> bool:
    """True si les deux chaînes contiennent le même nombre (à 0.001 près).
    Gère le cas décimal KELIA (0.03) vs affichage FPP (3 %)."""
    if not a or not b:
        return False
    try:
        def extract(s: str) -> tuple[float, bool]:
            has_pct = "%" in s
            clean = s.replace(",", ".").replace(" ", "").replace("%", "").strip()
            m = re.search(r"[-+]?\d+\.?\d*", clean)
            if not m:
                raise ValueError
            return float(m.group()), has_pct

        na, a_pct = extract(a)
        nb, b_pct = extract(b)

        # Décimal KELIA (ex: 0.03) face à un % FPP (ex: 3 %)
        if a_pct and not b_pct and 0.0 <= nb <= 1.0:
            nb *= 100
        elif b_pct and not a_pct and 0.0 <= na <= 1.0:
            na *= 100

        return abs(na - nb) < 0.001
    except (ValueError, AttributeError):
        return False


# ── 1. Versions FPP ───────────────────────────────────────────────────────────

def get_fpp_versions(db: Session, product_id: int) -> list[dict]:
    versions = (
        db.query(Version)
        .filter(Version.product_id == product_id, Version.artifact_type == "FicheDirect")
        .order_by(Version.version_number.desc())
        .all()
    )
    result = []
    for v in versions:
        item_count = (
            db.query(sqlfunc.count(FicheDirectItem.id))
            .filter(
                FicheDirectItem.product_id == product_id,
                FicheDirectItem.version_number == v.version_number,
            )
            .scalar() or 0
        )
        filled = (
            db.query(sqlfunc.count(FicheDirectItem.id))
            .filter(
                FicheDirectItem.product_id == product_id,
                FicheDirectItem.version_number == v.version_number,
                FicheDirectItem.status != "Information manquante",
            )
            .scalar() or 0
        )
        result.append(
            {
                "version": v.version_number,
                "label": v.version_label or f"V{v.version_number}",
                "created_at": v.created_at.isoformat() if v.created_at else None,
                "item_count": item_count,
                "filled_count": filled,
            }
        )
    return result


# ── 2. Non-régression (pur DB) ────────────────────────────────────────────────

def compare_fpp_versions(
    db: Session, product_id: int, v_ref: int, v_new: int
) -> dict:
    """
    Compare deux versions FPP pour détecter les régressions.

    Statuts retournés :
      stable     → valeur identique dans les deux versions
      modified   → valeur différente mais toutes deux renseignées
      regression → v_ref renseigné, v_new devient "Information manquante"
      added      → champ absent de v_ref, présent dans v_new
      removed    → champ présent dans v_ref, absent de v_new
    """
    ref_rows = (
        db.query(FicheDirectItem)
        .filter(
            FicheDirectItem.product_id == product_id,
            FicheDirectItem.version_number == v_ref,
        )
        .all()
    )
    new_rows = (
        db.query(FicheDirectItem)
        .filter(
            FicheDirectItem.product_id == product_id,
            FicheDirectItem.version_number == v_new,
        )
        .all()
    )

    ref_map: dict[tuple, FicheDirectItem] = {(r.sheet, r.parameter): r for r in ref_rows}
    new_map: dict[tuple, FicheDirectItem] = {(r.sheet, r.parameter): r for r in new_rows}
    all_keys = sorted(set(ref_map.keys()) | set(new_map.keys()), key=lambda x: (x[0] or "", x[1] or ""))

    items: list[dict] = []
    for sheet, param in all_keys:
        r = ref_map.get((sheet, param))
        n = new_map.get((sheet, param))
        v_ref_val = (r.value or NO_VAL) if r else None
        v_new_val = (n.value or NO_VAL) if n else None

        if r and n:
            if v_ref_val == v_new_val:
                status = "stable"
            elif _is_empty(v_new_val) and not _is_empty(v_ref_val):
                status = "regression"
            else:
                status = "modified"
        elif r:
            status = "removed"
        else:
            status = "added"

        items.append(
            {
                "sheet": sheet,
                "parameter": param,
                "v_ref_value": v_ref_val,
                "v_new_value": v_new_val,
                "status": status,
                "v_ref_status": r.status if r else None,
                "v_new_status": n.status if n else None,
                "section": (r or n).section if (r or n) else None,
            }
        )

    statuses = ("stable", "modified", "regression", "added", "removed")
    summary = {s: sum(1 for i in items if i["status"] == s) for s in statuses}
    return {"v_ref": v_ref, "v_new": v_new, "summary": summary, "items": items}


# ── 3. Parse fichier paramétrage KELIA ───────────────────────────────────────

# Colonnes candidates pour le nom du paramètre
_PARAM_HEADERS = {"parametre", "paramètre", "regle", "règle", "libelle", "libellé",
                  "nom", "champ", "field", "parameter", "rule", "rule_name",
                  "libelle parametre", "libellé paramètre", "libelle règle", "label"}
# Colonnes candidates pour la valeur
_VALUE_HEADERS = {"valeur", "value", "valeur kelia", "valeur parametre", "valeur actuelle",
                  "donnee", "donnée", "contenu", "setting", "rule_value",
                  "valeur saisie", "valeur renseignee", "valeur renseignée"}
# Colonnes candidates pour le produit
_PRODUIT_HEADERS = {"produit", "product", "code produit", "code_produit",
                    "nom produit", "nom_produit"}


def _detect_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    """Retourne (idx_param, idx_value, idx_produit) depuis une liste de headers."""
    h_norm = [_norm(h) for h in headers]

    def _find(keyword_set: set) -> int | None:
        # 1. Correspondance exacte
        idx = next((i for i, h in enumerate(h_norm) if h in keyword_set), None)
        if idx is not None:
            return idx
        # 2. Le header COMMENCE PAR un mot-clé (ex: "parametre si source" → "parametre")
        return next(
            (i for i, h in enumerate(h_norm)
             if any(h.startswith(kw) for kw in keyword_set if len(kw) >= 5)),
            None,
        )

    idx_param = _find(_PARAM_HEADERS)
    idx_val = _find(_VALUE_HEADERS)
    idx_produit = _find(_PRODUIT_HEADERS)
    # Fallback : prend les 2 premières colonnes non vides
    if idx_param is None and len(headers) >= 1:
        idx_param = 0
    if idx_val is None and len(headers) >= 2:
        idx_val = 1
    return idx_param, idx_val, idx_produit


def _cell_to_str(cell) -> str:
    """Convertit une cellule openpyxl en string, avec gestion du format %."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        fmt = getattr(cell, "number_format", None) or ""
        if "%" in fmt:
            pct = v * 100
            if pct == int(pct):
                return f"{int(pct)} %"
            return f"{pct:.2f}".replace(".", ",") + " %"
    return str(v).strip()


def _parse_excel(file_bytes: bytes) -> list[dict[str, str]]:
    """Parse un fichier Excel (.xlsx / .xls)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise ValueError("openpyxl requis pour lire les fichiers Excel")

    # data_only=True sans read_only pour accéder à number_format
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    best_rows: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows())
        if len(all_rows) < 2:
            continue

        # Cherche la ligne d'en-tête (première ligne non vide)
        header_idx = next(
            (i for i, r in enumerate(all_rows) if any(c.value is not None for c in r)), None
        )
        if header_idx is None:
            continue

        headers = [str(c.value).strip() if c.value is not None else "" for c in all_rows[header_idx]]
        idx_param, idx_val, idx_produit = _detect_columns(headers)
        if idx_param is None or idx_val is None:
            continue

        sheet_rows: list[dict[str, str]] = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c.value is None for c in row):
                continue
            param = _cell_to_str(row[idx_param]) if idx_param < len(row) else ""
            val = _cell_to_str(row[idx_val]) if idx_val < len(row) else ""
            produit = ""
            if idx_produit is not None and idx_produit < len(row):
                produit = _cell_to_str(row[idx_produit])
            if param and param.lower() not in ("none", "n/a", "#n/a"):
                sheet_rows.append({
                    "parametre_kelia": param,
                    "valeur_kelia": val or NO_VAL,
                    "produit_kelia": produit,
                })

        if len(sheet_rows) > len(best_rows):
            best_rows = sheet_rows

    return best_rows


def _parse_csv(file_bytes: bytes) -> list[dict[str, str]]:
    """Parse un fichier CSV (auto-détecte le séparateur)."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any(c.strip() for c in r)]
    if len(rows) < 2:
        return []

    headers = rows[0]
    idx_param, idx_val, idx_produit = _detect_columns(headers)
    result = []
    for row in rows[1:]:
        if len(row) <= max(idx_param or 0, idx_val or 0):
            continue
        param = row[idx_param].strip() if idx_param is not None else ""
        val = row[idx_val].strip() if idx_val is not None else ""
        produit = ""
        if idx_produit is not None and idx_produit < len(row):
            produit = row[idx_produit].strip()
        if param:
            result.append({
                "parametre_kelia": param,
                "valeur_kelia": val or NO_VAL,
                "produit_kelia": produit,
            })
    return result


def parse_parametrage_file(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    """
    Parse un fichier paramétrage KELIA (Excel ou CSV).
    Retourne une liste de { parametre_kelia, valeur_kelia, produit_kelia }.
    """
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext in ("xlsx", "xls", "xlsm"):
        rows = _parse_excel(file_bytes)
    elif ext == "csv":
        rows = _parse_csv(file_bytes)
    else:
        # Essaie Excel d'abord, puis CSV
        try:
            rows = _parse_excel(file_bytes)
        except Exception:
            rows = _parse_csv(file_bytes)

    if not rows:
        raise ValueError(
            f"Impossible de parser le fichier '{filename}'. "
            "Format attendu : Excel ou CSV avec colonnes Paramètre / Valeur."
        )
    return rows


# ── 4. LLM matching + comparaison ────────────────────────────────────────────

def _llm_match_and_compare(items_for_llm: list[dict], provider: str = "anthropic") -> list[dict]:
    """
    Pour chaque paramètre KELIA (avec ses candidats FPP pré-filtrés),
    le LLM choisit le meilleur match, compare les valeurs et explique.
    Retourne une liste d'objets {idx, parametre_fpp, sheet_fpp, valeur_fpp,
                                  statut, confiance_matching, explication}.
    """
    from app.services.ai_service import _call, set_active_provider, _sanitize_json
    import json as _json

    if not items_for_llm:
        return []

    set_active_provider(provider)

    prompt = (
        "Tu es expert paramétrage assurance-vie KELIA Art.83.\n\n"
        "Pour chaque paramètre KELIA ci-dessous, tu disposes d'une liste de candidats issus de la FPP "
        "(Fiche Produit Paramétré). Ta mission :\n"
        "1. Choisir LE meilleur candidat FPP (même concept, libellé peut différer).\n"
        "2. Comparer la valeur KELIA avec la valeur FPP.\n"
        "3. Attribuer un statut ET une confiance de matching.\n\n"
        "STATUTS :\n"
        '- "conforme"       : paramètre trouvé ET valeurs identiques ou numériquement équivalentes\n'
        '- "non_conforme"   : paramètre trouvé MAIS valeurs différentes (précise les deux dans explication)\n'
        '- "non_retrouve"   : aucun candidat FPP ne correspond au concept KELIA\n'
        '- "incertain"      : candidat trouvé mais tu doutes (libellé très différent / sens ambigu)\n\n'
        "CONFIANCE MATCHING :\n"
        '- "certain"   : correspondance claire et directe\n'
        '- "probable"  : bonne correspondance sémantique malgré libellé différent\n'
        '- "incertain" : tu hésites entre plusieurs candidats ou sens ambigu\n\n'
        "Réponds UNIQUEMENT avec un JSON array (même ordre que les paramètres reçus) :\n"
        '[{\n'
        '  "idx": <int>,\n'
        '  "parametre_fpp": "<libellé exact du candidat choisi, ou null>",\n'
        '  "sheet_fpp": "<onglet FPP ou null>",\n'
        '  "valeur_fpp": "<valeur FPP ou null>",\n'
        '  "statut": "<conforme|non_conforme|non_retrouve|incertain>",\n'
        '  "confiance_matching": "<certain|probable|incertain>",\n'
        '  "explication": "<1-2 phrases : pourquoi ce match + pourquoi conforme/non_conforme>"\n'
        '}]\n\n'
        f"Données :\n{_json.dumps(items_for_llm, ensure_ascii=False)}"
    )

    try:
        raw = _call("gpt-4o", prompt, max_tokens=2000)
        raw = _sanitize_json(raw)
        return _json.loads(raw)
    except Exception:
        # Fallback : premier candidat sans explication LLM
        return [
            {
                "idx": it["idx"],
                "parametre_fpp": it["candidats_fpp"][0]["parametre"] if it["candidats_fpp"] else None,
                "sheet_fpp": it["candidats_fpp"][0]["sheet"] if it["candidats_fpp"] else None,
                "valeur_fpp": it["candidats_fpp"][0]["valeur"] if it["candidats_fpp"] else None,
                "statut": "incertain" if it["candidats_fpp"] else "non_retrouve",
                "confiance_matching": "incertain",
                "explication": "Analyse IA indisponible — correspondance algorithmique.",
            }
            for it in items_for_llm
        ]


# ── 5. Matrice KELIA ↔ FPP ───────────────────────────────────────────────────

def compare_parametrage_vs_fpp(
    db: Session,
    product_id: int,
    fpp_version: int,
    kelia_rows: list[dict[str, str]],
    provider: str = "anthropic",
) -> dict:
    """
    Compare le paramétrage KELIA uploadé (kelia_rows) avec la FPP du produit.
    Retourne UNIQUEMENT les lignes issues du fichier KELIA (pas d'absent_kelia).

    Statuts :
      conforme     → trouvé + valeurs identiques / équivalentes
      non_conforme → trouvé + valeurs différentes
      non_retrouve → aucun paramètre FPP correspondant
      incertain    → match trouvé mais confiance faible
    """
    fpp_rows = (
        db.query(FicheDirectItem)
        .filter(
            FicheDirectItem.product_id == product_id,
            FicheDirectItem.version_number == fpp_version,
        )
        .all()
    )

    # Index FPP dédupliqué : préfère les entrées avec valeur non vide
    fpp_by_norm: dict[str, FicheDirectItem] = {}
    for r in fpp_rows:
        key = _norm(r.parameter)
        if not key:
            continue
        existing = fpp_by_norm.get(key)
        if existing is None or (_is_empty(existing.value) and not _is_empty(r.value)):
            fpp_by_norm[key] = r

    # Pour chaque paramètre KELIA, calcule top-5 candidats FPP
    items_for_llm: list[dict] = []
    for idx, krow in enumerate(kelia_rows):
        k_norm = _norm(krow["parametre_kelia"])
        scored: list[tuple[float, FicheDirectItem]] = []

        for fkey, frow in fpp_by_norm.items():
            if fkey == k_norm:
                sc = 1.0
            elif k_norm and (k_norm in fkey or fkey in k_norm) and len(min(k_norm, fkey, key=len)) >= 6:
                sc = 0.85
            else:
                sc = _word_sim(k_norm, fkey)
            if sc > 0.3:
                scored.append((sc, frow))

        scored.sort(key=lambda x: -x[0])
        candidats = [
            {
                "parametre": frow.parameter,
                "valeur": frow.value or NO_VAL,
                "sheet": frow.sheet or "",
                "score": round(sc, 2),
            }
            for sc, frow in scored[:5]
        ]

        items_for_llm.append({
            "idx": idx,
            "parametre_kelia": krow["parametre_kelia"],
            "valeur_kelia": krow["valeur_kelia"],
            "candidats_fpp": candidats,
        })

    # LLM : matching sémantique + comparaison valeurs + explication
    llm_results = _llm_match_and_compare(items_for_llm, provider)
    llm_by_idx: dict[int, dict] = {r.get("idx", i): r for i, r in enumerate(llm_results)}

    # Post-correction : override confiance_matching avec le score algorithmique
    # Le LLM marque souvent "incertain" sur des libellés différents mais sémantiquement identiques.
    for it_llm in items_for_llm:
        idx = it_llm["idx"]
        llm = llm_by_idx.get(idx)
        if not llm or not llm.get("parametre_fpp"):
            continue

        chosen_norm = _norm(llm["parametre_fpp"])
        k_norm = _norm(it_llm["parametre_kelia"])

        if chosen_norm == k_norm:
            algo_sc = 1.0
        elif k_norm and (k_norm in chosen_norm or chosen_norm in k_norm) and len(min(k_norm, chosen_norm, key=len)) >= 6:
            algo_sc = 0.85
        else:
            algo_sc = _word_sim(k_norm, chosen_norm)

        # Élève la confiance si le score algo le justifie
        cur = llm.get("confiance_matching", "incertain")
        if algo_sc >= 0.75 and cur in ("incertain", "probable"):
            llm["confiance_matching"] = "certain"
        elif algo_sc >= 0.5 and cur == "incertain":
            llm["confiance_matching"] = "probable"

        # Si le statut LLM est "incertain" mais le matching est maintenant certain/probable :
        # réévalue le statut en comparant les valeurs algorithmiquement
        if llm.get("statut") == "incertain" and llm["confiance_matching"] in ("certain", "probable"):
            k_val = it_llm.get("valeur_kelia", "")
            fpp_val = llm.get("valeur_fpp") or ""
            if _is_empty(fpp_val):
                llm["statut"] = "non_retrouve"
            elif _norm(k_val) == _norm(fpp_val) or _numeric_equiv(k_val, fpp_val):
                llm["statut"] = "conforme"
            else:
                llm["statut"] = "non_conforme"

    # Construction des items de sortie (1 ligne = 1 ligne KELIA)
    items: list[dict] = []
    for idx, krow in enumerate(kelia_rows):
        llm = llm_by_idx.get(idx, {})
        statut = llm.get("statut", "non_retrouve")
        if statut not in ("conforme", "non_conforme", "non_retrouve", "incertain"):
            statut = "non_retrouve"
        items.append(
            {
                "produit_kelia": krow.get("produit_kelia", ""),
                "parametre_kelia": krow["parametre_kelia"],
                "valeur_kelia": krow["valeur_kelia"],
                "parametre_fpp": llm.get("parametre_fpp"),
                "valeur_fpp": llm.get("valeur_fpp"),
                "sheet": llm.get("sheet_fpp"),
                "status": statut,
                "confiance_matching": llm.get("confiance_matching", "incertain"),
                "explication": llm.get("explication", ""),
                "trouve_dans_fpp": statut not in ("non_retrouve",),
            }
        )

    statuses = ("conforme", "non_conforme", "non_retrouve", "incertain")
    summary = {s: sum(1 for i in items if i["status"] == s) for s in statuses}

    conforme_count = summary["conforme"]
    non_conforme_count = summary["non_conforme"]
    total = len(items)
    taux = round(
        conforme_count / max(1, total - summary["non_retrouve"]) * 100
    )

    if non_conforme_count == 0 and summary["non_retrouve"] == 0 and summary["incertain"] == 0:
        statut_global = "Conforme"
    elif non_conforme_count / max(1, total) < 0.2:
        statut_global = "Conforme avec réserves"
    else:
        statut_global = "Non conforme"

    # Stats par produit
    products: dict[str, dict] = {}
    for item in items:
        prod_key = item.get("produit_kelia", "") or ""
        if prod_key not in products:
            products[prod_key] = {
                "total": 0, "conforme": 0, "non_conforme": 0,
                "non_retrouve": 0, "incertain": 0, "taux_conformite": 0,
            }
        s = item["status"]
        products[prod_key]["total"] += 1
        if s in products[prod_key]:
            products[prod_key][s] += 1

    for pstats in products.values():
        pstats["taux_conformite"] = round(
            pstats["conforme"] / max(1, pstats["conforme"] + pstats["non_conforme"]) * 100
        )

    return {
        "fpp_version": fpp_version,
        "kelia_rows": len(kelia_rows),
        "fpp_rows": len(fpp_rows),
        "summary": summary,
        "taux_conformite": taux,
        "statut_global": statut_global,
        "products": products,
        "items": items,
    }


# ── 6. Historisation ──────────────────────────────────────────────────────────

def save_recette_history(
    db: Session,
    product_id: int,
    fpp_version: int,
    filename_kelia: str,
    provider: str,
    result: dict,
) -> int:
    """Sauvegarde une recette en historique. Retourne l'id créé."""
    import json as _json
    from app.models.recette_history import RecetteHistory
    from datetime import datetime

    entry = RecetteHistory(
        product_id=product_id,
        fpp_version=fpp_version,
        filename_kelia=filename_kelia,
        provider=provider,
        created_at=datetime.utcnow(),
        kelia_rows=result.get("kelia_rows", 0),
        taux_conformite=result.get("taux_conformite", 0),
        statut_global=result.get("statut_global", ""),
        result_json=_json.dumps(result, ensure_ascii=False),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry.id


def get_recette_history_list(db: Session, product_id: int) -> list[dict]:
    """Retourne la liste légère des recettes historisées pour un produit."""
    from app.models.recette_history import RecetteHistory

    entries = (
        db.query(RecetteHistory)
        .filter(RecetteHistory.product_id == product_id)
        .order_by(RecetteHistory.created_at.desc())
        .all()
    )
    return [
        {
            "id": e.id,
            "fpp_version": e.fpp_version,
            "filename_kelia": e.filename_kelia,
            "provider": e.provider,
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "kelia_rows": e.kelia_rows,
            "taux_conformite": e.taux_conformite,
            "statut_global": e.statut_global,
        }
        for e in entries
    ]


def get_recette_history_detail(db: Session, history_id: int) -> dict | None:
    """Retourne le résultat complet d'une recette historisée."""
    import json as _json
    from app.models.recette_history import RecetteHistory

    entry = db.query(RecetteHistory).filter(RecetteHistory.id == history_id).first()
    if not entry:
        return None
    result = _json.loads(entry.result_json)
    result["history_id"] = entry.id
    result["filename_kelia"] = entry.filename_kelia
    result["created_at"] = entry.created_at.isoformat() if entry.created_at else None
    result["annotations"] = _json.loads(entry.annotations_json) if entry.annotations_json else {}
    return result


def export_recette_history_excel(db: Session, history_id: int) -> bytes:
    """Génère un fichier Excel pour une recette historisée."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json as _json
    from app.models.recette_history import RecetteHistory

    entry = db.query(RecetteHistory).filter(RecetteHistory.id == history_id).first()
    if not entry:
        raise ValueError(f"Historique {history_id} introuvable")

    result = _json.loads(entry.result_json)
    items = result.get("items", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recette Paramétrage"

    # Styles
    header_fill = PatternFill("solid", fgColor="A100FF")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    STATUS_COLORS = {
        "conforme":     "D1FAE5",
        "non_conforme": "FEE2E2",
        "non_retrouve": "FEF3C7",
        "incertain":    "F3E8FF",
    }

    # En-tête
    headers = ["Paramètre KELIA", "Valeur KELIA", "Paramètre FPP", "Valeur FPP", "Statut", "Confiance", "Explication IA"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 22

    # Données
    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("parametre_kelia", ""),
            item.get("valeur_kelia", ""),
            item.get("parametre_fpp", ""),
            item.get("valeur_fpp", ""),
            item.get("status", ""),
            item.get("confiance_matching", ""),
            item.get("explication", ""),
        ]
        status = item.get("status", "")
        fill_color = STATUS_COLORS.get(status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))
            if col <= 6:
                cell.fill = row_fill

    # Onglet métadonnées
    ws_meta = wb.create_sheet("Informations")
    meta = [
        ("Fichier KELIA", entry.filename_kelia),
        ("Version FPP", entry.fpp_version),
        ("Modèle IA", entry.provider),
        ("Date", entry.created_at.strftime("%d/%m/%Y %H:%M") if entry.created_at else ""),
        ("Taux conformité", f"{entry.taux_conformite}%"),
        ("Lignes KELIA", entry.kelia_rows),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=str(v))

    # Largeurs colonnes
    col_widths = [35, 18, 35, 18, 16, 14, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 7. Analyse non-régression : 2 fichiers paramétrage ───────────────────────

def _llm_nonreg_compare(items_for_llm: list[dict], provider: str = "openai-gpt5") -> list[dict]:
    """
    LLM compare chaque paramètre du Fichier 1 avec ses candidats du Fichier 2.
    Retourne : {idx, parametre_f2, valeur_f2, statut, confiance_matching, explication}
    statuts : stable | modifie | supprime
    """
    from app.services.ai_service import _call, set_active_provider, _sanitize_json
    import json as _json

    if not items_for_llm:
        return []

    set_active_provider(provider)

    prompt = (
        "Tu es expert paramétrage assurance-vie KELIA Art.83.\n\n"
        "Tu analyses l'évolution entre deux versions de fichiers paramétrage KELIA.\n"
        "Pour chaque paramètre du Fichier 1 (référence), tu as des candidats du Fichier 2 (nouvelle version).\n\n"
        "STATUTS :\n"
        '- "stable"   : même concept, valeurs identiques ou numériquement équivalentes\n'
        '- "modifie"  : même concept trouvé, mais la valeur a changé\n'
        '- "supprime" : aucun candidat Fichier 2 ne correspond à ce paramètre\n\n'
        "CONFIANCE MATCHING :\n"
        '- "certain"   : libellé identique ou quasi-identique\n'
        '- "probable"  : même concept, libellé légèrement différent\n'
        '- "incertain" : tu hésites\n\n'
        "Réponds UNIQUEMENT en JSON array (même ordre que les paramètres reçus) :\n"
        '[{\n'
        '  "idx": <int>,\n'
        '  "parametre_f2": "<libellé exact du candidat Fichier 2, ou null si supprimé>",\n'
        '  "valeur_f2": "<valeur Fichier 2, ou null>",\n'
        '  "statut": "<stable|modifie|supprime>",\n'
        '  "confiance_matching": "<certain|probable|incertain>",\n'
        '  "explication": "<1-2 phrases : matching + comparaison valeurs>"\n'
        '}]\n\n'
        f"Données :\n{_json.dumps(items_for_llm, ensure_ascii=False)}"
    )

    try:
        raw = _call("gpt-4o", prompt, max_tokens=2000)
        raw = _sanitize_json(raw)
        return _json.loads(raw)
    except Exception:
        return [
            {
                "idx": it["idx"],
                "parametre_f2": it["candidats_f2"][0]["parametre"] if it["candidats_f2"] else None,
                "valeur_f2": it["candidats_f2"][0]["valeur"] if it["candidats_f2"] else None,
                "statut": "incertain" if it["candidats_f2"] else "supprime",
                "confiance_matching": "incertain",
                "explication": "Analyse IA indisponible.",
            }
            for it in items_for_llm
        ]


def compare_two_parametrage_files(
    file1_rows: list[dict[str, str]],
    file2_rows: list[dict[str, str]],
    filename1: str = "Fichier 1",
    filename2: str = "Fichier 2",
    provider: str = "openai-gpt5",
) -> dict:
    """
    Compare deux fichiers paramétrage KELIA.
    file1 = référence, file2 = nouvelle version.
    Retourne une liste d'items avec statut : stable / modifie / supprime / ajoute.
    """
    # Index file2 par nom normalisé
    f2_by_norm: dict[str, dict] = {}
    for row in file2_rows:
        key = _norm(row["parametre_kelia"])
        if key:
            f2_by_norm[key] = row

    # Pour chaque param file1, trouver top-5 candidats dans file2
    items_for_llm: list[dict] = []
    for idx, row1 in enumerate(file1_rows):
        k_norm = _norm(row1["parametre_kelia"])
        scored: list[tuple[float, dict]] = []

        for f2key, f2row in f2_by_norm.items():
            if f2key == k_norm:
                sc = 1.0
            elif k_norm and (k_norm in f2key or f2key in k_norm) and len(min(k_norm, f2key, key=len)) >= 6:
                sc = 0.85
            else:
                sc = _word_sim(k_norm, f2key)
            if sc > 0.3:
                scored.append((sc, f2row))

        scored.sort(key=lambda x: -x[0])
        candidats = [
            {"parametre": r["parametre_kelia"], "valeur": r["valeur_kelia"], "score": round(sc, 2)}
            for sc, r in scored[:5]
        ]

        items_for_llm.append({
            "idx": idx,
            "parametre_f1": row1["parametre_kelia"],
            "valeur_f1": row1["valeur_kelia"],
            "candidats_f2": candidats,
        })

    # LLM matching
    llm_results = _llm_nonreg_compare(items_for_llm, provider)
    llm_by_idx: dict[int, dict] = {r.get("idx", i): r for i, r in enumerate(llm_results)}

    # Post-correction confiance (même logique que compare_parametrage_vs_fpp)
    for it_llm in items_for_llm:
        idx = it_llm["idx"]
        llm = llm_by_idx.get(idx)
        if not llm or not llm.get("parametre_f2"):
            continue
        chosen_norm = _norm(llm["parametre_f2"])
        k_norm = _norm(it_llm["parametre_f1"])
        if chosen_norm == k_norm:
            algo_sc = 1.0
        elif k_norm and (k_norm in chosen_norm or chosen_norm in k_norm) and len(min(k_norm, chosen_norm, key=len)) >= 6:
            algo_sc = 0.85
        else:
            algo_sc = _word_sim(k_norm, chosen_norm)
        cur = llm.get("confiance_matching", "incertain")
        if algo_sc >= 0.75 and cur in ("incertain", "probable"):
            llm["confiance_matching"] = "certain"
        elif algo_sc >= 0.5 and cur == "incertain":
            llm["confiance_matching"] = "probable"
        # Réévalue statut si "incertain" mais confiance certaine
        if llm.get("statut") == "incertain" and llm["confiance_matching"] in ("certain", "probable"):
            v1 = it_llm.get("valeur_f1", "")
            v2 = llm.get("valeur_f2") or ""
            if _is_empty(v2):
                llm["statut"] = "supprime"
            elif _norm(v1) == _norm(v2) or _numeric_equiv(v1, v2):
                llm["statut"] = "stable"
            else:
                llm["statut"] = "modifie"

    # Params file2 non matchés → "ajoute"
    matched_f2_norms: set[str] = set()
    items: list[dict] = []

    for idx, row1 in enumerate(file1_rows):
        llm = llm_by_idx.get(idx, {})
        statut = llm.get("statut", "supprime")
        if statut not in ("stable", "modifie", "supprime"):
            statut = "supprime"
        if llm.get("parametre_f2"):
            matched_f2_norms.add(_norm(llm["parametre_f2"]))
        items.append({
            "parametre_f1": row1["parametre_kelia"],
            "valeur_f1": row1["valeur_kelia"],
            "parametre_f2": llm.get("parametre_f2"),
            "valeur_f2": llm.get("valeur_f2"),
            "status": statut,
            "confiance_matching": llm.get("confiance_matching", "incertain"),
            "explication": llm.get("explication", ""),
        })

    for row2 in file2_rows:
        if _norm(row2["parametre_kelia"]) not in matched_f2_norms:
            items.append({
                "parametre_f1": None,
                "valeur_f1": None,
                "parametre_f2": row2["parametre_kelia"],
                "valeur_f2": row2["valeur_kelia"],
                "status": "ajoute",
                "confiance_matching": "certain",
                "explication": f"Paramètre présent uniquement dans {filename2}.",
            })

    statuses = ("stable", "modifie", "supprime", "ajoute")
    summary = {s: sum(1 for i in items if i["status"] == s) for s in statuses}

    return {
        "filename1": filename1,
        "filename2": filename2,
        "f1_rows": len(file1_rows),
        "f2_rows": len(file2_rows),
        "summary": summary,
        "items": items,
    }
