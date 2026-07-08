"""
Service de non-régression entre deux fichiers paramétrage KELIA.

Architecture :
  1. Parse les deux fichiers (tous les onglets Excel, colonnes param/valeur/commentaire)
  2. Agent Expert Comparaison (LLM) — matching sémantique, analyse métier assurance Art.83
  3. Agent Synthèse (LLM) — régressions priorisées, recommandations
  4. Fallback algorithmique si le LLM échoue
"""
from __future__ import annotations
import io, json, csv, re, unicodedata
from typing import Any


# ── Normalisation ─────────────────────────────────────────────────────────────

def _norm(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip().rstrip("*").strip()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


NO_VAL = "Information manquante"
_EMPTY_VALS = {NO_VAL, "", None, "none", "n/a", "#n/a", "#valeur!"}

_PARAM_HEADERS = {
    "parametre", "paramètre", "regle", "règle", "libelle", "libellé",
    "nom", "champ", "field", "parameter", "rule", "rule name",
    "libelle parametre", "libellé paramètre", "label",
    "si source", "si_source", "designation", "désignation",
    "parametre kelia", "nom du parametre", "nom parametre",
    "libelle regle", "libelle de la regle", "intitule", "intitulé",
    "nom regle", "code regle", "code parametre",
}
_VALUE_HEADERS = {
    "valeur", "value", "valeur kelia", "valeur parametre", "valeur actuelle",
    "donnee", "donnée", "contenu", "setting", "rule value",
    "valeur saisie", "valeur renseignee", "valeur renseignée",
    "valeur du parametre", "valeur de la regle",
}
_COMMENT_HEADERS = {
    "commentaire", "commentaires", "comment", "remarque", "remarques",
    "note", "notes", "precision", "précision", "observation", "observations",
    "regle de gestion", "règle de gestion", "rg", "description",
    "complement", "complément", "information complementaire",
}

# Mots-clés partiels pour détection souple (ex : "Paramètre KELIA", "Nom du paramètre")
_PARAM_KW   = ("param", "regle", "libelle", "source", "intitul", "designation", "nom regle", "nom param")
_VALUE_KW   = ("valeur", "value", "donnee", "contenu", "saisie", "renseign")
_COMMENT_KW = ("comment", "remarque", "note", "precision", "observation", "complement")


def _detect_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    h_norm = [_norm(h) for h in headers]

    # Phase 1 : correspondance exacte
    idx_param   = next((i for i, h in enumerate(h_norm) if h in _PARAM_HEADERS), None)
    idx_val     = next((i for i, h in enumerate(h_norm) if h in _VALUE_HEADERS), None)
    idx_comment = next((i for i, h in enumerate(h_norm) if h in _COMMENT_HEADERS), None)

    # Phase 2 : correspondance partielle (ex : "Paramètre KELIA", "Valeur du paramètre")
    if idx_param is None:
        idx_param = next(
            (i for i, h in enumerate(h_norm) if h and any(kw in h for kw in _PARAM_KW)),
            None,
        )
    if idx_val is None:
        idx_val = next(
            (i for i, h in enumerate(h_norm) if h and i != idx_param and any(kw in h for kw in _VALUE_KW)),
            None,
        )
    if idx_comment is None:
        idx_comment = next(
            (i for i, h in enumerate(h_norm)
             if h and i not in (idx_param, idx_val) and any(kw in h for kw in _COMMENT_KW)),
            None,
        )

    # Phase 3 : fallback sur les premières colonnes NON VIDES
    # (évite de tomber sur les colonnes A/B vides ou avec numéros de ligne)
    non_empty = [i for i, h in enumerate(h_norm) if h]

    if idx_param is None:
        idx_param = non_empty[0] if non_empty else 0
    if idx_val is None:
        cands = [i for i in non_empty if i != idx_param]
        idx_val = cands[0] if cands else (idx_param + 1 if idx_param is not None else 1)
    if idx_comment is None:
        cands = [i for i in non_empty if i not in (idx_param, idx_val)]
        idx_comment = cands[0] if cands else None

    return idx_param, idx_val, idx_comment


# ── Parsing ───────────────────────────────────────────────────────────────────

def _find_best_header_row(rows: list, max_scan: int = 15) -> int:
    """
    Retourne l'index de la ligne d'en-têtes la plus probable
    parmi les max_scan premières lignes.
    Stratégie : score = nb colonnes non vides + bonus si contient des mots-clés connus.
    """
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:max_scan]):
        if not row or all(c is None for c in row):
            continue
        cells = [_norm(str(c)) if c is not None else "" for c in row]
        score = sum(1 for h in cells if h)  # bonus pour chaque colonne non vide
        if any(h in _PARAM_HEADERS or any(kw in h for kw in _PARAM_KW) for h in cells if h):
            score += 5
        if any(h in _VALUE_HEADERS or any(kw in h for kw in _VALUE_KW) for h in cells if h):
            score += 3
        if any(h in _COMMENT_HEADERS or any(kw in h for kw in _COMMENT_KW) for h in cells if h):
            score += 1
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _col_looks_like_params(data_rows: list, idx: int, sample: int = 8) -> bool:
    """
    Vérifie que la colonne idx contient des libellés (chaînes non numériques, longueur > 3).
    Retourne True si c'est le cas pour au moins 2 lignes sur l'échantillon.
    """
    hits = 0
    for row in data_rows[:sample]:
        if not row or idx >= len(row) or row[idx] is None:
            continue
        v = str(row[idx]).strip()
        if len(v) > 3 and not v.replace(".", "").replace(",", "").replace("%", "").replace(" ", "").isdigit():
            hits += 1
    return hits >= 2


def _parse_excel_all_sheets(file_bytes: bytes) -> list[dict[str, str]]:
    """
    Lit TOUS les onglets Excel.
    Retourne {parametre, valeur, commentaire}.
    Détection automatique des colonnes avec fallback C/D/E (format KELIA standard).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    all_rows: list[dict[str, str]] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Trouver la meilleure ligne d'en-têtes (parmi les 15 premières)
        header_idx = _find_best_header_row(rows)
        headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        idx_param, idx_val, idx_comment = _detect_columns(headers)

        data_rows = rows[header_idx + 1:]

        # Si la colonne param détectée ne ressemble pas à des libellés,
        # essayer les colonnes C/D/E (indices 2/3/4) — format KELIA standard
        if not _col_looks_like_params(data_rows, idx_param or 0):
            if len(headers) >= 5:
                # Essai 1 : colonnes C/D/E
                if _col_looks_like_params(data_rows, 2):
                    idx_param, idx_val, idx_comment = 2, 3, 4 if len(headers) > 4 else None
                # Essai 2 : scanner toutes les colonnes pour trouver une colonne de libellés
                else:
                    for candidate in range(len(headers)):
                        if _col_looks_like_params(data_rows, candidate):
                            idx_param = candidate
                            idx_val   = candidate + 1 if candidate + 1 < len(headers) else None
                            idx_comment = candidate + 2 if candidate + 2 < len(headers) else None
                            break

        if idx_param is None or idx_val is None:
            continue

        for row in data_rows:
            if not row or all(c is None for c in row):
                continue
            param = str(row[idx_param]).strip() if idx_param < len(row) and row[idx_param] is not None else ""
            val   = str(row[idx_val]).strip()   if idx_val   < len(row) and row[idx_val]   is not None else ""
            comment = ""
            if idx_comment is not None and idx_comment < len(row) and row[idx_comment] is not None:
                comment = str(row[idx_comment]).strip()
            if not param or _norm(param) in _EMPTY_VALS:
                continue
            key = _norm(param)
            if key not in seen:
                all_rows.append({
                    "parametre":   param,
                    "valeur":      val or NO_VAL,
                    "commentaire": comment,
                })
                seen.add(key)

    return all_rows


def _parse_csv_file(file_bytes: bytes) -> list[dict[str, str]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any(c.strip() for c in r)]
    if len(rows) < 2:
        return []
    headers = rows[0]
    idx_param, idx_val, idx_comment = _detect_columns(headers)
    result: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if len(row) <= max(idx_param or 0, idx_val or 0):
            continue
        param   = row[idx_param].strip() if idx_param is not None else ""
        val     = row[idx_val].strip()   if idx_val   is not None else ""
        comment = row[idx_comment].strip() if (idx_comment is not None and idx_comment < len(row)) else ""
        if not param or _norm(param) in _EMPTY_VALS:
            continue
        key = _norm(param)
        if key not in seen:
            result.append({
                "parametre":   param,
                "valeur":      val or NO_VAL,
                "commentaire": comment,
            })
            seen.add(key)
    return result


def parse_file(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    """Retourne une liste de {parametre, valeur, commentaire} depuis un Excel ou CSV."""
    from app.services.recette_service import parse_parametrage_file
    rows = parse_parametrage_file(file_bytes, filename)
    return [
        {
            "parametre":   r["parametre_kelia"],
            "valeur":      r["valeur_kelia"],
            "commentaire": "",
        }
        for r in rows
    ]


# ── Agent Expert Comparaison ──────────────────────────────────────────────────

def _compare_llm_expert(
    rows_v1: list[dict], rows_v2: list[dict],
    v1_name: str, v2_name: str, provider: str,
) -> dict:
    """
    Agent Expert Comparaison.
    LLM positionné comme expert recette assurance-vie Art. 83 —
    matching sémantique + analyse métier + 4 niveaux de criticité.
    """
    from app.services.ai_service import _call, set_active_provider, _parse_json
    set_active_provider(provider)

    # Limiter à 80 rows par version pour rester dans les limites token LLM
    MAX_ROWS = 80
    r1 = rows_v1[:MAX_ROWS]
    r2 = rows_v2[:MAX_ROWS]

    prompt = f"""Tu es un expert senior en recette de paramétrage assurance-vie (Art. 83 retraite collective).
Tu maîtrises parfaitement :
- Le paramétrage des produits assurance et épargne retraite collective
- Les règles de gestion : frais (acquisition, gestion, arrérages, transfert), garanties, taux, provisions
- Les référentiels produits KELIA et les impacts fonctionnels des écarts de paramétrage
- La migration et reprise de données en assurance collective
- Les contrôles de non-régression en recette fonctionnelle

MISSION : Comparer les deux versions du paramétrage KELIA et identifier TOUTES les régressions.

=== VERSION V1 (référence) : {v1_name} ({len(rows_v1)} paramètres au total) ===
{json.dumps(r1, ensure_ascii=False)}

=== VERSION V2 (nouvelle version) : {v2_name} ({len(rows_v2)} paramètres au total) ===
{json.dumps(r2, ensure_ascii=False)}

RAISONNEMENT EN 4 ÉTAPES (expert BMAD) :

ÉTAPE 1 — STRUCTURE
Identifie le champ SI_SOURCE et les regroupements métier (frais, garanties, taux, conditions de sortie, options...).

ÉTAPE 2 — MATCHING SÉMANTIQUE
Fais correspondre les paramètres même si leur libellé diffère :
- "Frais sur versements" ≈ "Frais sur cotisations" ≈ "Frais d'entrée" selon contexte
- "Frais de gestion encours" ≈ "Frais gestion en cours" ≈ "Frais annuels de gestion"
- Détecte les regroupements ou découpages de paramètres entre versions

ÉTAPE 3 — ANALYSE MÉTIER DES VALEURS
- Changements de valeurs numériques (ex : 3% → 5% = écart tarifaire majeur)
- Changements d'unités : "0,50% annuel" ≠ "0,50% mensuel" → ÉCART CRITIQUE sur calcul
- Changements de règles de gestion ou conditions
- Incohérences entre plusieurs paramètres liés (ex : frais contradictoires)

ÉTAPE 4 — ANALYSE DES COMMENTAIRES (champ "commentaire")
- Reformulation purement rédactionnelle → conforme
- Évolution documentaire sans impact métier → noter en remarques
- Modification avec impact fonctionnel ou réglementaire → statut ÉCART même si valeur identique

DÉFINITION DES STATUTS :
- "conforme"  : paramètre présent dans V1 et V2, valeur identique, commentaires cohérents
- "ecart"     : présent dans les 2 versions, valeur différente OU commentaire avec impact métier
- "nouveau"   : présent en V2 uniquement (ajouté)
- "supprime"  : présent en V1 uniquement (supprimé)

NIVEAUX DE CRITICITÉ (obligatoire pour CHAQUE paramètre) :
- "critique" : impact direct sur calcul prestations, frais réglementaires, conditions légales,
               ou risque de reprise de données erronée (migration)
- "élevé"    : impact sur règles de gestion importantes, conditions tarifaires, garanties
- "moyen"    : paramètres fonctionnels secondaires, impact indirect
- "faible"   : libellés cosmétiques, reformulations sans impact métier

Retourne UNIQUEMENT ce JSON valide (sans texte avant ou après) :
{{
  "taux_conformite": <entier 0-100>,
  "controles_realises": <nombre total de paramètres distincts analysés>,
  "items": [
    {{
      "parametre": "<libellé canonique du paramètre>",
      "valeur_v1": "<valeur en V1 ou null>",
      "valeur_v2": "<valeur en V2 ou null>",
      "commentaire_v1": "<commentaire V1 ou null si vide>",
      "commentaire_v2": "<commentaire V2 ou null si vide>",
      "status": "conforme|ecart|nouveau|supprime",
      "criticite": "critique|élevé|moyen|faible",
      "explication": "<explication métier 1-2 phrases, null pour conforme>",
      "impact_metier": "<impact fonctionnel précis sur le produit assurance, null pour conforme>",
      "remarques": "<observations sur les commentaires ou le contexte, null si rien>",
      "recommandation": "<action corrective concrète recommandée, null pour conforme>"
    }}
  ]
}}"""

    raw = _call("gpt-4o", prompt, max_tokens=16000)
    return _parse_json(raw)


# ── Agent Synthèse ────────────────────────────────────────────────────────────

def _synthesize_llm(
    items: list[dict],
    v1_name: str, v2_name: str,
    taux_conformite: int, provider: str,
) -> dict:
    """
    Agent Synthèse : génère synthèse métier, régressions priorisées, recommandations.
    """
    from app.services.ai_service import _call, _sanitize_json

    non_conformes = [i for i in items if i.get("status") != "conforme"]
    if not non_conformes:
        return {
            "synthese": (
                f"Analyse complète de {v1_name} vers {v2_name} : aucun écart détecté. "
                "Les deux versions du paramétrage KELIA sont identiques à 100%. "
                "Aucune régression n'a été identifiée."
            ),
            "regressions_prioritaires": [],
            "recommandations": ["Aucune action corrective requise — les deux versions sont conformes."],
        }

    critiques = [i for i in non_conformes if i.get("criticite") in ("critique", "élevé")]
    top_n = min(5, len(non_conformes))

    prompt = f"""Tu es expert senior en recette de paramétrage assurance-vie Art. 83.

Analyse de non-régression : {v1_name} → {v2_name}
Taux de conformité global : {taux_conformite}%
Écarts détectés : {len(non_conformes)} (dont {len(critiques)} de criticité critique/élevé)

Écarts et évolutions identifiés :
{json.dumps(non_conformes, ensure_ascii=False, indent=2)}

En tant qu'expert recette, génère :

1. Une SYNTHÈSE MÉTIER (3-4 phrases) qui :
   - Met en évidence les régressions majeures
   - Évalue l'impact sur le produit assurance-vie et les assurés
   - Indique le niveau de risque global pour la mise en production

2. La liste des {top_n} RÉGRESSIONS LES PLUS IMPORTANTES à traiter en priorité
   (format : "PARAMÈTRE : nature précise de l'écart et son impact")

3. Des RECOMMANDATIONS CONCRÈTES (3-5 actions) pour corriger ou investiguer les écarts

Retourne UNIQUEMENT ce JSON valide :
{{
  "synthese": "<synthèse métier 3-4 phrases>",
  "regressions_prioritaires": [
    "<Paramètre X : écart sur la valeur Y, impact sur Z>",
    "..."
  ],
  "recommandations": [
    "<Action concrète 1>",
    "..."
  ]
}}"""

    raw = _call("gpt-4o", prompt, max_tokens=2000)
    return json.loads(_sanitize_json(raw))


# ── Fallback algorithmique ────────────────────────────────────────────────────

def _compare_algo(
    rows_v1: list[dict], rows_v2: list[dict],
    v1_name: str, v2_name: str,
) -> dict:
    """Diff algorithmique (fallback si LLM indisponible)."""
    map_v1 = {_norm(r["parametre"]): r for r in rows_v1}
    map_v2 = {_norm(r["parametre"]): r for r in rows_v2}
    all_keys = sorted(set(map_v1.keys()) | set(map_v2.keys()))

    items = []
    for key in all_keys:
        r1 = map_v1.get(key)
        r2 = map_v2.get(key)
        label   = (r1 or r2)["parametre"]
        v1_val  = r1["valeur"]      if r1 else None
        v2_val  = r2["valeur"]      if r2 else None
        c1      = r1.get("commentaire", "") if r1 else ""
        c2      = r2.get("commentaire", "") if r2 else ""

        if r1 and r2:
            status = "conforme" if _norm(v1_val or "") == _norm(v2_val or "") else "ecart"
        elif r1:
            status = "supprime"
        else:
            status = "nouveau"

        items.append({
            "parametre":      label,
            "valeur_v1":      v1_val,
            "valeur_v2":      v2_val,
            "commentaire_v1": c1 or None,
            "commentaire_v2": c2 or None,
            "status":         status,
            "criticite":      "moyen",
            "explication":    None,
            "impact_metier":  None,
            "remarques":      None,
            "recommandation": None,
        })

    n_conf = sum(1 for i in items if i["status"] == "conforme")
    taux   = round(n_conf / max(1, len(items)) * 100)
    non_conformes = len(items) - n_conf

    return {
        "taux_conformite":        taux,
        "controles_realises":     len(items),
        "items":                  items,
        "synthese":               (
            f"Comparaison algorithmique (sans IA) : {len(items)} paramètres analysés, "
            f"{non_conformes} écart(s) détecté(s)."
        ),
        "regressions_prioritaires": [],
        "recommandations": [
            "Mode algorithmique — LLM indisponible. "
            "Relancer avec un modèle IA pour une analyse sémantique complète."
        ],
    }


# ── Point d'entrée principal ──────────────────────────────────────────────────

def compare_parametrage_files(
    v1_bytes: bytes, v1_name: str,
    v2_bytes: bytes, v2_name: str,
    provider: str = "anthropic",
) -> dict:
    """
    Compare deux fichiers paramétrage KELIA.
    Moteur principal : LLM expert assurance (matching sémantique + analyse métier).
    Fallback algorithmique si le LLM échoue.
    """
    rows_v1 = parse_file(v1_bytes, v1_name)
    rows_v2 = parse_file(v2_bytes, v2_name)

    # Avertissement si les deux fichiers parsés sont identiques
    norms_v1 = {_norm(r["parametre"]): _norm(r["valeur"]) for r in rows_v1}
    norms_v2 = {_norm(r["parametre"]): _norm(r["valeur"]) for r in rows_v2}
    parsing_warning = None
    if norms_v1 == norms_v2 and len(norms_v1) > 0:
        parsing_warning = (
            "⚠️ Les deux fichiers ont été parsés de façon identique. "
            f"Exemples extraits : {[r['parametre'] for r in rows_v1[:3]]}. "
            "Utilisez 'Vérifier le parsing' pour contrôler les colonnes détectées."
        )

    try:
        comparison = _compare_llm_expert(rows_v1, rows_v2, v1_name, v2_name, provider)
        items     = comparison.get("items", [])
        taux      = comparison.get("taux_conformite", 0)
        controles = comparison.get("controles_realises", len(items))

        try:
            synthesis = _synthesize_llm(items, v1_name, v2_name, taux, provider)
        except Exception:
            # Synthèse non critique — on continue sans
            synthesis = {
                "synthese":                "",
                "regressions_prioritaires": [],
                "recommandations":          [],
            }

    except Exception as e:
        fallback  = _compare_algo(rows_v1, rows_v2, v1_name, v2_name)
        fallback["synthese"] = f"[Mode algorithmique — LLM indisponible : {e}] " + fallback["synthese"]
        items     = fallback["items"]
        taux      = fallback["taux_conformite"]
        controles = fallback["controles_realises"]
        synthesis = {
            "synthese":                fallback["synthese"],
            "regressions_prioritaires": fallback["regressions_prioritaires"],
            "recommandations":          fallback["recommandations"],
        }

    statuses = ("conforme", "ecart", "nouveau", "supprime")
    summary  = {s: sum(1 for i in items if i.get("status") == s) for s in statuses}

    criticites = ("critique", "élevé", "moyen", "faible")
    non_conf   = [i for i in items if i.get("status") != "conforme"]
    criticite_summary = {
        c: sum(1 for i in non_conf if i.get("criticite") == c) for c in criticites
    }

    return {
        "v1_name":                 v1_name,
        "v2_name":                 v2_name,
        "v1_count":                len(rows_v1),
        "v2_count":                len(rows_v2),
        "controles_realises":      controles,
        "taux_conformite":         taux,
        "synthese":                synthesis.get("synthese", ""),
        "regressions_prioritaires": synthesis.get("regressions_prioritaires", []),
        "recommandations":         synthesis.get("recommandations", []),
        "summary":                 summary,
        "criticite_summary":       criticite_summary,
        "parsing_warning":         parsing_warning,
        "items":                   items,
    }


# ── Historisation ─────────────────────────────────────────────────────────────

def save_nonreg_history(
    db, product_id: int, filename_v1: str, filename_v2: str, provider: str, result: dict
) -> int:
    import json as _json
    from app.models.nonreg_history import NonRegHistory

    stable = result.get("summary", {}).get("conforme", 0)
    total = result.get("controles_realises", 1) or 1
    entry = NonRegHistory(
        product_id=product_id,
        filename_v1=filename_v1,
        filename_v2=filename_v2,
        provider=provider,
        v1_rows=result.get("v1_count", 0),
        v2_rows=result.get("v2_count", 0),
        taux_stable=round(stable / total * 100, 1),
        result_json=_json.dumps(result, ensure_ascii=False),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry.id


def get_nonreg_history_list(db, product_id: int) -> list[dict]:
    from app.models.nonreg_history import NonRegHistory
    rows = (
        db.query(NonRegHistory)
        .filter(NonRegHistory.product_id == product_id)
        .order_by(NonRegHistory.created_at.desc())
        .all()
    )
    return [
        {
            "id": r.id,
            "filename_v1": r.filename_v1,
            "filename_v2": r.filename_v2,
            "provider": r.provider,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "v1_rows": r.v1_rows,
            "v2_rows": r.v2_rows,
            "taux_stable": r.taux_stable,
        }
        for r in rows
    ]


def get_nonreg_history_detail(db, history_id: int) -> dict | None:
    import json as _json
    from app.models.nonreg_history import NonRegHistory

    entry = db.query(NonRegHistory).filter(NonRegHistory.id == history_id).first()
    if not entry:
        return None
    result = _json.loads(entry.result_json)
    result["history_id"] = entry.id
    result["filename_v1"] = entry.filename_v1
    result["filename_v2"] = entry.filename_v2
    result["created_at"] = entry.created_at.isoformat() if entry.created_at else None
    result["annotations"] = _json.loads(entry.annotations_json) if entry.annotations_json else {}
    return result


def export_nonreg_history_excel(db, history_id: int) -> bytes:
    import io, openpyxl, json as _json
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.models.nonreg_history import NonRegHistory

    entry = db.query(NonRegHistory).filter(NonRegHistory.id == history_id).first()
    if not entry:
        raise ValueError(f"Historique {history_id} introuvable")

    result = _json.loads(entry.result_json)
    items = result.get("items", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non-Régression"

    header_fill = PatternFill("solid", fgColor="A100FF")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    STATUS_COLORS = {
        "conforme":  "D1FAE5",
        "ecart":     "FEE2E2",
        "nouveau":   "E0E7FF",
        "supprime":  "FEF3C7",
    }

    headers = ["Paramètre", "Valeur V1", "Valeur V2", "Statut", "Criticité", "Analyse IA"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(items, 2):
        status = item.get("status", "")
        row_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        row_data = [
            item.get("parametre", ""),
            item.get("valeur_v1", ""),
            item.get("valeur_v2", ""),
            status,
            item.get("criticite", ""),
            item.get("explication", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            if col <= 5:
                cell.fill = row_fill

    ws_meta = wb.create_sheet("Informations")
    meta = [
        ("Fichier V1", entry.filename_v1),
        ("Fichier V2", entry.filename_v2),
        ("Modèle IA", entry.provider),
        ("Date", entry.created_at.strftime("%d/%m/%Y %H:%M") if entry.created_at else ""),
        ("Taux stable", f"{entry.taux_stable}%"),
        ("Lignes V1", entry.v1_rows),
        ("Lignes V2", entry.v2_rows),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=str(v))

    col_widths = [40, 20, 20, 16, 12, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
